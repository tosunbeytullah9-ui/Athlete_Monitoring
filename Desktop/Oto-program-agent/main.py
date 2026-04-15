# -*- coding: utf-8 -*-
"""
Exercise Library — main application entry point.
Run with:  uvicorn main:app --reload
"""

from __future__ import annotations

import io
import json
import os
import re
import secrets
import unicodedata
from typing import List, Optional
from datetime import datetime, timedelta

from fastapi import Depends, FastAPI, HTTPException, Query, Request, status
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from jose import JWTError, jwt
from passlib.context import CryptContext
from pydantic import BaseModel
from sqlalchemy import Column, Float, Integer, String, Text, create_engine, inspect, or_, text
from sqlalchemy.orm import DeclarativeBase, Session, sessionmaker

# ── Database ──────────────────────────────────────────────────────────────────
os.makedirs("data", exist_ok=True)
DATABASE_URL = "sqlite:///./data/exercises.db"
engine = create_engine(DATABASE_URL, connect_args={"check_same_thread": False})
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)


class Base(DeclarativeBase):
    pass


class ExerciseDB(Base):
    __tablename__ = "exercises"

    id               = Column(Integer, primary_key=True, index=True)
    name             = Column(String,  nullable=False)
    name_tr          = Column(String,  nullable=True)
    movement_pattern = Column(String,  nullable=False)
    category         = Column(String,  nullable=True)
    equipment        = Column(Text,    default="[]")
    difficulty       = Column(String,  nullable=False)
    training_quality = Column(String,  nullable=True, default="Strength")
    primary_muscles  = Column(Text,    default="[]")
    secondary_muscles= Column(Text,    default="[]")
    description      = Column(Text,    nullable=True)
    coaching_cues    = Column(Text,    default="[]")
    sport_tags       = Column(Text,    default="[]")
    gender_tags      = Column(Text,    default='["all"]')
    video_url        = Column(String,  nullable=True)
    notes            = Column(Text,    nullable=True)


class ProgramDB(Base):
    __tablename__ = "programs"

    id                 = Column(Integer, primary_key=True, index=True)
    name               = Column(String,  nullable=False)
    start_date         = Column(String,  nullable=True)
    duration_weeks     = Column(Integer, nullable=False, default=8)
    days_per_week      = Column(Integer, nullable=False, default=3)
    goal               = Column(String,  nullable=False, default="strength")
    periodization_type = Column(String,  nullable=False, default="block")
    athlete_level      = Column(String,  nullable=True)
    sport              = Column(String,  nullable=True)
    athlete_id         = Column(Integer, nullable=True)
    team_id            = Column(Integer, nullable=True)
    notes              = Column(Text,    nullable=True)
    created_at         = Column(String,  nullable=False)


class ProgramWeekDB(Base):
    __tablename__ = "program_weeks"

    id          = Column(Integer, primary_key=True, index=True)
    program_id  = Column(Integer, nullable=False)
    week_number = Column(Integer, nullable=False)
    phase       = Column(String,  nullable=True)
    notes       = Column(Text,    nullable=True)


class ProgramSessionDB(Base):
    __tablename__ = "program_sessions"

    id           = Column(Integer, primary_key=True, index=True)
    week_id      = Column(Integer, nullable=False)
    day_number   = Column(Integer, nullable=False)
    session_name = Column(String,  nullable=True)
    sort_order   = Column(Integer, nullable=False, default=0)
    notes        = Column(Text,    nullable=True)


class OneRMDB(Base):
    __tablename__ = "one_rm_records"

    id            = Column(Integer, primary_key=True, index=True)
    exercise_id   = Column(Integer, nullable=True)
    exercise_name = Column(String,  nullable=False)
    weight_kg     = Column(String,  nullable=False)
    test_date     = Column(String,  nullable=True)
    notes         = Column(Text,    nullable=True)
    created_at    = Column(String,  nullable=False)
    athlete_id    = Column(Integer, nullable=True)


class ExerciseRMRelationshipDB(Base):
    __tablename__ = "rm_relationships"

    id            = Column(Integer, primary_key=True, index=True)
    exercise_name = Column(String,  nullable=False)   # derived exercise
    source_name   = Column(String,  nullable=False)   # base exercise (must have 1RM)
    factor        = Column(Float,   nullable=False)   # multiplier (e.g. 0.75)
    per_hand      = Column(Integer, nullable=False, default=0)  # 1 = DB/single-arm → divide by 2


# ── Relationship seed data (from iSPORTSi Exercise Menu) ───────────────────────
# (exercise_name, source_name, factor, per_hand)
RM_RELATIONSHIPS_SEED = [
    # ── Olympic / Power — HANG CLEAN based ────────────────────────────────────
    ("DB HANG CLEAN",                   "HANG CLEAN", 0.60,  1),
    ("BLOCK CLEAN",                     "HANG CLEAN", 0.80,  0),
    ("BLOCK SNATCH",                    "HANG CLEAN", 0.525, 0),
    ("BLOCK CLEAN & JERK",              "HANG CLEAN", 0.75,  0),
    ("CLEAN/FRONT SQUAT COMBO",         "HANG CLEAN", 0.75,  0),
    ("HANG CLEAN/PUSH JERK",            "HANG CLEAN", 0.75,  0),
    ("SNATCH/SQUAT/JERK",               "HANG CLEAN", 0.60,  0),
    ("HANG SNATCH",                     "HANG CLEAN", 0.60,  0),
    ("DB SNATCH",                       "HANG SNATCH",1.10,  1),
    ("TRAP BAR SQUAT JUMP",             "HANG CLEAN", 0.22,  0),
    ("DB BOX JUMP",                     "HANG CLEAN", 0.20,  1),
    ("DB JUMP SQUAT",                   "HANG CLEAN", 0.20,  1),
    ("DB JUMP SQUAT + BOX JUMP",        "HANG CLEAN", 0.20,  1),
    ("DB SPLIT JUMP",                   "HANG CLEAN", 0.20,  1),
    ("VERTIMAX W/ CATCH +WEIGHT",       "HANG CLEAN", 0.20,  1),
    ("VERTIMAX CONT +WEIGHT",           "HANG CLEAN", 0.20,  1),
    # ── Jerk based ────────────────────────────────────────────────────────────
    ("FRONT PUSH JERK",                 "PUSH JERK",  0.85,  0),
    ("PUSH PRESS",                      "PUSH JERK",  0.80,  0),
    ("DB JERK",                         "PUSH JERK",  0.70,  1),
    ("ALT. ARM JAMMER",                 "PUSH JERK",  1.15,  0),
    ("DOUBLE ARM JAMMER",               "PUSH JERK",  1.10,  0),
    # ── Squat based ───────────────────────────────────────────────────────────
    ("FRONT SQUAT",                     "SQUAT",      0.75,  0),
    ("SAFETY BAR SQUAT",                "SQUAT",      0.85,  0),
    ("BELT SQUAT",                      "SQUAT",      0.80,  0),
    ("OVERHEAD SQUAT",                  "SQUAT",      0.35,  0),
    ("PAUSE SQUAT",                     "SQUAT",      0.80,  0),
    ("LATERAL SQUAT",                   "SQUAT",      0.35,  0),
    ("GOBLET SQUAT",                    "SQUAT",      0.35,  0),
    ("GOBLET LATERAL SQUAT",            "SQUAT",      0.15,  0),
    ("ECCENTRIC GOBLET SQUAT",          "SQUAT",      0.35,  0),
    ("PAUSE GOBLET SQUAT",              "SQUAT",      0.30,  0),
    ("SINGLE LEG DB BENCH SQUAT",       "SQUAT",      0.30,  1),
    ("SINBLE LEG SB BENCH SQUAT",       "SQUAT",      0.25,  1),
    ("SINBLE LEG BB BENCH SQUAT",       "SQUAT",      0.60,  1),
    ("DB STABILITY BALL WALL SQUAT",    "SQUAT",      0.45,  1),
    ("BB SPLIT SQUAT",                  "SQUAT",      0.50,  0),
    ("DB SPLIT SQUAT",                  "SQUAT",      0.30,  1),
    ("BB STEP-UP",                      "SQUAT",      0.475, 0),
    ("GOBLET STEP-UP",                  "SQUAT",      0.22,  0),
    ("LATERAL DB STEP-UP",              "SQUAT",      0.25,  1),
    ("CROSSOVER DB STEP UP",            "SQUAT",      0.25,  1),
    ("DB STEP-UP",                      "SQUAT",      0.35,  1),
    ("BB LUNGE",                        "SQUAT",      0.48,  0),
    ("DB LUNGE OR REVERSE LUNGE",       "SQUAT",      0.35,  1),
    ("LATERAL LUNGE/SQUAT",             "SQUAT",      0.38,  0),
    ("3-WAY LUNGE",                     "SQUAT",      0.38,  0),
    ("GOBLET LATERAL LUNGE",            "SQUAT",      0.35,  1),
    ("SUMO DEADLIFT",                   "SQUAT",      0.85,  0),
    ("TRAP BAR DEADLIFT",               "SQUAT",      1.20,  0),
    ("SL TRAP BAR DEADLIFT",            "SQUAT",      0.58,  0),
    ("REVERSE HYPER",                   "SQUAT",      0.50,  0),
    ("SL REVERSE HYPER",                "SQUAT",      0.30,  1),
    ("BB HIP BRIDGE",                   "SQUAT",      0.90,  0),
    # ── Hip extension — HANG CLEAN based ──────────────────────────────────────
    ("RDL",                             "HANG CLEAN", 1.20,  0),
    ("BB GOODMORNING",                  "HANG CLEAN", 0.70,  0),
    ("SL DB RDL",                       "HANG CLEAN", 0.70,  1),
    ("SL BB RDL",                       "HANG CLEAN", 0.70,  0),
    ("SL DEADLIFT",                     "HANG CLEAN", 0.70,  0),
    ("DB SWING",                        "HANG CLEAN", 0.40,  1),
    ("SA SUITCASE DEADLIFT",            "HANG CLEAN", 0.50,  0),
    ("STAGGERED DEADLIFT",              "HANG CLEAN", 0.85,  0),
    ("SL LANDMINE RDL",                 "HANG CLEAN", 0.85,  1),
    ("DEADLIFT",                        "HANG CLEAN", 1.20,  0),
    # ── Bench based ───────────────────────────────────────────────────────────
    ("PAUSE BENCH PRESS",               "BENCH",      0.90,  0),
    ("1 BOARD PRESS",                   "BENCH",      0.95,  0),
    ("2 BOARD PRESS",                   "BENCH",      1.00,  0),
    ("3 BOARD PRESS",                   "BENCH",      1.05,  0),
    ("4 BOARD PRESS",                   "BENCH",      1.10,  0),
    ("FLOOR PRESS",                     "BENCH",      0.925, 0),
    ("DB FLOOR PRESS",                  "BENCH",      0.75,  1),
    ("CLOSE GRIP BENCH",                "BENCH",      0.90,  0),
    ("DB BENCH",                        "BENCH",      0.80,  1),
    ("SB DB BENCH PRESS",               "BENCH",      0.80,  1),
    ("THICK BAR BENCH PRESS",           "BENCH",      0.90,  0),
    ("INCLINE",                         "BENCH",      0.75,  0),
    ("CLOSE GRIP INCLINE BENCH",        "BENCH",      0.75,  0),
    ("DB INCLINE",                      "BENCH",      0.70,  1),
    ("MILITARY",                        "BENCH",      0.55,  0),
    ("DB MILITARY",                     "BENCH",      0.55,  1),
    ("ALT DB PRESS",                    "BENCH",      0.55,  1),
    ("K BOARD DB PRESS",                "BENCH",      0.50,  1),
    ("SL DB MILITARY",                  "BENCH",      0.50,  1),
    ("THICK BAR MILITARY",              "BENCH",      0.60,  0),
    ("DB ROW",                          "BENCH",      0.75,  1),
    ("THICK GRIP DB ROW",               "BENCH",      0.50,  1),
    ("SWISS BAR ROW",                   "BENCH",      0.65,  0),
    ("PG LANDMINE ROW",                 "BENCH",      0.65,  0),
    ("SA LANDMINE ROW",                 "BENCH",      0.95,  1),
    ("SA BARBELL ROW",                  "BENCH",      0.35,  0),
    ("HAMMER ROW",                      "BENCH",      1.00,  0),
    ("BB INCLINE PRESS",                "BENCH",      0.75,  0),
    # ── BB Incline Press based ────────────────────────────────────────────────
    ("CLOSE GRIP INCLINE PRESS",        "BB INCLINE PRESS", 0.86, 0),
    ("SWISS BAR INCLINE PRESS",         "BB INCLINE PRESS", 0.85, 0),
    ("THICK BAR INCLINE PRESS",         "BB INCLINE PRESS", 0.96, 0),
    ("VIKING PRESS",                    "BB INCLINE PRESS", 1.06, 0),
    ("SL DB MILITARY PRESS",            "BB INCLINE PRESS", 0.66, 1),
    ("DB CURL TO PRESS",                "BB INCLINE PRESS", 0.53, 1),
    ("THICK GRIP DB CURL TO PRESS",     "BB INCLINE PRESS", 0.53, 1),
    ("LUNGE STANCE ALT. DB PRESS",      "BB INCLINE PRESS", 0.60, 1),
    ("TALL KNEELING ALT. DB PRESS",     "BB INCLINE PRESS", 0.56, 1),
    ("DB STEP UP & PRESS",              "BB INCLINE PRESS", 0.73, 1),
    ("THICK GRIP DB MILITARY PRESS",    "BB INCLINE PRESS", 0.60, 1),
    ("THICK BAR MILITARY PRESS",        "BB INCLINE PRESS", 0.70, 0),
    ("ECCENTRIC DB INCLINE BENCH PRESS","BB INCLINE PRESS", 0.93, 1),
    ("DB INCLINE BENCH PRESS",          "BB INCLINE PRESS", 0.93, 1),
    ("ALT LOCKOUT DB INCLINE BENCH PRESS","BB INCLINE PRESS",0.80,1),
    # ── Horizontal pull — HANG CLEAN based ────────────────────────────────────
    ("BARBELL ROW",                     "HANG CLEAN", 0.75,  0),
    ("THICK BAR ROW",                   "HANG CLEAN", 0.65,  0),
    ("TWO-ARM DB ROW",                  "HANG CLEAN", 0.70,  1),
    ("FB POSITION DB ROW",              "HANG CLEAN", 0.65,  1),
    ("UNDERGRIP ROW",                   "HANG CLEAN", 0.70,  0),
    ("BB SHRUG",                        "HANG CLEAN", 1.00,  0),
]


class SessionExerciseDB(Base):
    __tablename__ = "session_exercises"

    id              = Column(Integer, primary_key=True, index=True)
    session_id      = Column(Integer, nullable=False)
    exercise_id     = Column(Integer, nullable=True)
    exercise_name   = Column(String,  nullable=True)
    sort_order      = Column(Integer, nullable=False, default=0)
    sets            = Column(Integer, nullable=True)
    reps            = Column(String,  nullable=True)   # default reps for all sets
    intensity_type  = Column(String,  nullable=True)   # legacy / default intensity type
    intensity_value = Column(String,  nullable=True)   # legacy / default intensity value
    rest_seconds    = Column(Integer, nullable=True)
    tempo           = Column(String,  nullable=True)
    notes           = Column(Text,    nullable=True)
    pairing_group   = Column(String,  nullable=True)  # "A", "B", "C" …
    pairing_slot    = Column(Integer, nullable=True)   # 1 veya 2


class SessionExerciseSetDB(Base):
    """Per-set prescription: each set of an exercise can have its own
    reps, intensity type (weight_kg / pct_1rm / rpe / rir) and value."""

    __tablename__ = "session_exercise_sets"

    id                  = Column(Integer, primary_key=True, index=True)
    session_exercise_id = Column(Integer, nullable=False)
    set_number          = Column(Integer, nullable=False)   # 1-based
    reps                = Column(String,  nullable=True)    # overrides exercise default
    intensity_type      = Column(String,  nullable=True)    # pct_1rm | rpe | rir | load_kg | bodyweight
    intensity_value     = Column(String,  nullable=True)
    notes               = Column(Text,    nullable=True)


class UserDB(Base):
    __tablename__ = "users"

    id            = Column(Integer, primary_key=True, index=True)
    username      = Column(String,  nullable=False, unique=True)
    password_hash = Column(String,  nullable=True)
    role          = Column(String,  nullable=False, default="athlete")  # "coach" or "athlete"
    created_at    = Column(String,  nullable=False)
    google_id     = Column(String,  nullable=True, unique=True)
    email         = Column(String,  nullable=True)


class AthleteDB(Base):
    __tablename__ = "athletes"

    id         = Column(Integer, primary_key=True, index=True)
    user_id    = Column(Integer, nullable=True)    # optional link to a login account
    name       = Column(String,  nullable=False)
    birth_year = Column(Integer, nullable=True)
    sport      = Column(String,  nullable=True)    # gymnastics, football, general
    position   = Column(String,  nullable=True)    # event / playing position
    gender     = Column(String,  nullable=True)    # male, female
    level      = Column(String,  nullable=True)    # beginner, intermediate, advanced, elite
    notes      = Column(Text,    nullable=True)
    created_at = Column(String,  nullable=False)


class TeamDB(Base):
    __tablename__ = "teams"

    id         = Column(Integer, primary_key=True, index=True)
    name       = Column(String,  nullable=False)
    sport      = Column(String,  nullable=True)
    notes      = Column(Text,    nullable=True)
    created_at = Column(String,  nullable=False)


class TeamMemberDB(Base):
    __tablename__ = "team_members"

    id         = Column(Integer, primary_key=True, index=True)
    team_id    = Column(Integer, nullable=False)
    athlete_id = Column(Integer, nullable=False)


class MovementPatternDB(Base):
    __tablename__ = "movement_patterns"

    id    = Column(Integer, primary_key=True, index=True)
    name  = Column(String,  nullable=False, unique=True)
    color = Column(String,  nullable=True)   # hex color, e.g. "#FF5733"


# ── Gymnastics Competition Models ─────────────────────────────────────────────

class GuestAthleteDB(Base):
    __tablename__ = "guest_athletes"

    id         = Column(Integer, primary_key=True, index=True)
    name       = Column(String,  nullable=False)
    country    = Column(String,  nullable=True)
    gender     = Column(String,  nullable=False)   # "male" or "female"
    birth_year = Column(Integer, nullable=True)
    notes      = Column(Text,    nullable=True)
    created_at = Column(String,  nullable=False)


class CompetitionEventDB(Base):
    __tablename__ = "competition_events"

    id                  = Column(Integer, primary_key=True, index=True)
    name                = Column(String,  nullable=False)
    date                = Column(String,  nullable=False)   # YYYY-MM-DD
    location            = Column(String,  nullable=True)
    level               = Column(String,  nullable=True)    # club/national/international
    notes               = Column(Text,    nullable=True)
    created_by_coach_id = Column(Integer, nullable=True)
    created_at          = Column(String,  nullable=False)


class CompetitionResultDB(Base):
    __tablename__ = "competition_results"

    id                   = Column(Integer, primary_key=True, index=True)
    competition_event_id = Column(Integer, nullable=False)
    athlete_id           = Column(Integer, nullable=True)    # FK to athletes
    guest_athlete_id     = Column(Integer, nullable=True)    # FK to guest_athletes
    apparatus            = Column(String,  nullable=False)   # e.g. "VT", "FX"
    d_score              = Column(Float,   nullable=False, default=0.0)
    e_score              = Column(Float,   nullable=False, default=0.0)
    penalty              = Column(Float,   nullable=False, default=0.0)
    bonus_point          = Column(Float,   nullable=False, default=0.0)
    d_score_2            = Column(Float,   nullable=True)    # Vault jump 2
    e_score_2            = Column(Float,   nullable=True)
    penalty_2            = Column(Float,   nullable=True)
    bonus_point_2        = Column(Float,   nullable=True)
    final_score          = Column(Float,   nullable=True)    # d + e - penalty + bonus
    rank                 = Column(Integer, nullable=True)
    notes                = Column(Text,    nullable=True)
    created_at           = Column(String,  nullable=False)


Base.metadata.create_all(bind=engine)

# ── Safe migrations ────────────────────────────────────────────────────────────
def _migrate():
    with engine.connect() as conn:
        for stmt in [
            "ALTER TABLE exercises ADD COLUMN training_quality TEXT DEFAULT 'Strength'",
            "ALTER TABLE programs ADD COLUMN athlete_id INTEGER",
            "ALTER TABLE one_rm_records ADD COLUMN athlete_id INTEGER",
            "ALTER TABLE programs ADD COLUMN team_id INTEGER",
            "ALTER TABLE users ADD COLUMN google_id TEXT",
            "ALTER TABLE users ADD COLUMN email TEXT",
            "ALTER TABLE session_exercises ADD COLUMN pairing_group TEXT",
            "ALTER TABLE session_exercises ADD COLUMN pairing_slot INTEGER",
            "ALTER TABLE competition_results ADD COLUMN bonus_point FLOAT DEFAULT 0.0",
            "ALTER TABLE competition_results ADD COLUMN d_score_2 FLOAT",
            "ALTER TABLE competition_results ADD COLUMN e_score_2 FLOAT",
            "ALTER TABLE competition_results ADD COLUMN penalty_2 FLOAT",
            "ALTER TABLE competition_results ADD COLUMN bonus_point_2 FLOAT",
        ]:
            try:
                conn.execute(text(stmt))
                conn.commit()
            except Exception:
                pass


_migrate()

# ── Auth ───────────────────────────────────────────────────────────────────────
_KEY_FILE = "data/secret.key"
# Priority: 1. SECRET_KEY env var  2. persistent file  3. generate new
SECRET_KEY = os.environ.get("SECRET_KEY", "").strip()
if not SECRET_KEY:
    if os.path.exists(_KEY_FILE):
        with open(_KEY_FILE) as _f:
            SECRET_KEY = _f.read().strip()
    if not SECRET_KEY:
        SECRET_KEY = secrets.token_hex(32)
        try:
            with open(_KEY_FILE, "w") as _f:
                _f.write(SECRET_KEY)
        except OSError:
            pass  # read-only filesystem — key lives in memory this boot

ALGORITHM   = "HS256"
TOKEN_HOURS = 12

GOOGLE_CLIENT_ID = os.environ.get("GOOGLE_CLIENT_ID", "").strip()

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")


def _hash(password: str) -> str:
    return pwd_context.hash(password)


def _verify(password: str, hashed: str) -> bool:
    return pwd_context.verify(password, hashed)


def _create_token(user_id: int, role: str) -> str:
    exp = datetime.utcnow() + timedelta(hours=TOKEN_HOURS)
    return jwt.encode({"sub": str(user_id), "role": role, "exp": exp}, SECRET_KEY, algorithm=ALGORITHM)


def _decode_token(token: str) -> dict:
    return jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])


# ── Pydantic schemas ───────────────────────────────────────────────────────────

class ExerciseIn(BaseModel):
    name:              str
    name_tr:           Optional[str]       = None
    movement_pattern:  str
    category:          Optional[str]       = None
    equipment:         Optional[List[str]] = []
    difficulty:        str
    training_quality:  Optional[str]       = "Strength"
    primary_muscles:   Optional[List[str]] = []
    secondary_muscles: Optional[List[str]] = []
    description:       Optional[str]       = None
    coaching_cues:     Optional[List[str]] = []
    sport_tags:        Optional[List[str]] = []
    gender_tags:       Optional[List[str]] = ["all"]
    video_url:         Optional[str]       = None
    notes:             Optional[str]       = None


class ProgramIn(BaseModel):
    name:               str
    start_date:         Optional[str]  = None
    duration_weeks:     int            = 8
    days_per_week:      int            = 3
    goal:               str            = "strength"
    periodization_type: str            = "block"
    athlete_level:      Optional[str]  = None
    sport:              Optional[str]  = None
    notes:              Optional[str]  = None
    athlete_id:         Optional[int]  = None
    team_id:            Optional[int]  = None

class SessionIn(BaseModel):
    week_number:  int
    day_number:   int
    session_name: Optional[str] = None
    notes:        Optional[str] = None

class SessionUpdateIn(BaseModel):
    session_name: Optional[str] = None
    notes:        Optional[str] = None

class ExerciseSetIn(BaseModel):
    set_number:      int
    reps:            Optional[str] = None
    intensity_type:  Optional[str] = None
    intensity_value: Optional[str] = None
    notes:           Optional[str] = None

class ExerciseSetUpdateIn(BaseModel):
    reps:            Optional[str] = None
    intensity_type:  Optional[str] = None
    intensity_value: Optional[str] = None
    notes:           Optional[str] = None

class SessionExerciseIn(BaseModel):
    exercise_id:     Optional[int] = None
    exercise_name:   Optional[str] = None
    sets:            Optional[int] = None
    reps:            Optional[str] = None
    intensity_type:  Optional[str] = None   # kept for backward compat
    intensity_value: Optional[str] = None   # kept for backward compat
    rest_seconds:    Optional[int] = None
    tempo:           Optional[str] = None
    notes:           Optional[str] = None
    sets_data:       Optional[List[ExerciseSetIn]] = None
    pairing_group:   Optional[str] = None   # "A", "B", "C" …
    pairing_slot:    Optional[int] = None   # 1 veya 2

class WeekPhaseIn(BaseModel):
    phase: str

class OneRMIn(BaseModel):
    exercise_id:   Optional[int] = None
    exercise_name: str
    weight_kg:     str
    test_date:     Optional[str] = None
    notes:         Optional[str] = None
    athlete_id:    Optional[int] = None

class LoginIn(BaseModel):
    username: str
    password: str

class SignupIn(BaseModel):
    username: str
    password: str
    role:     str = "athlete"

class GoogleLoginIn(BaseModel):
    credential: str  # Google ID token

class UserCreateIn(BaseModel):
    username: str
    password: str
    role:     str = "athlete"  # "coach" or "athlete"

class AthleteIn(BaseModel):
    name:       str
    password:   Optional[str]  = None   # auto-creates login account on new; updates pw on edit
    birth_year: Optional[int]  = None
    sport:      Optional[str]  = None
    position:   Optional[str]  = None
    gender:     Optional[str]  = None
    level:      Optional[str]  = None
    notes:      Optional[str]  = None

class TeamIn(BaseModel):
    name:  str
    sport: Optional[str] = None
    notes: Optional[str] = None

class TeamMemberIn(BaseModel):
    athlete_id: int

class GuestAthleteIn(BaseModel):
    name:       str
    country:    Optional[str] = None
    gender:     str                     # "male" or "female"
    birth_year: Optional[int] = None
    notes:      Optional[str] = None

class CompetitionEventIn(BaseModel):
    name:     str
    date:     str                       # YYYY-MM-DD
    location: Optional[str] = None
    level:    Optional[str] = None      # club/national/international
    notes:    Optional[str] = None

class CompetitionResultIn(BaseModel):
    competition_event_id: int
    athlete_id:           Optional[int]   = None
    guest_athlete_id:     Optional[int]   = None
    apparatus:            str
    d_score:              float           = 0.0
    e_score:              float           = 0.0
    penalty:              float           = 0.0
    bonus_point:          float           = 0.0
    d_score_2:            Optional[float] = None
    e_score_2:            Optional[float] = None
    penalty_2:            Optional[float] = None
    bonus_point_2:        Optional[float] = None
    rank:                 Optional[int]   = None
    notes:                Optional[str]   = None

# ── Helpers ────────────────────────────────────────────────────────────────────

def _athlete_username(name: str, db: Session, exclude_id: int = None) -> str:
    """Derive a unique login username from an athlete's full name.
    Turkish characters are transliterated; spaces and special chars become dots.
    e.g.  'Ahmet Yılmaz' → 'ahmet.yilmaz'  (or 'ahmet.yilmaz2' if taken)
    """
    tr = str.maketrans("ıİğĞüÜşŞöÖçÇ", "iigguussoocc")
    base = name.lower().translate(tr)
    base = re.sub(r"[^a-z0-9]+", ".", base).strip(".")
    candidate, n = base, 1
    while True:
        q = db.query(UserDB).filter(UserDB.username == candidate)
        existing = q.first()
        if not existing:
            break
        # If this username already belongs to the athlete being updated, keep it
        if exclude_id and existing.id == exclude_id:
            break
        candidate = f"{base}{n}"
        n += 1
    return candidate


def _j(v):
    return json.dumps(v, ensure_ascii=False) if v is not None else "[]"

def _u(v):
    try:
        return json.loads(v) if v else []
    except Exception:
        return []

def _row(r: ExerciseDB) -> dict:
    return {
        "id":               r.id,
        "name":             r.name,
        "name_tr":          r.name_tr,
        "movement_pattern": r.movement_pattern,
        "category":         r.category,
        "equipment":        _u(r.equipment),
        "difficulty":       r.difficulty,
        "training_quality": r.training_quality or "Strength",
        "primary_muscles":  _u(r.primary_muscles),
        "secondary_muscles":_u(r.secondary_muscles),
        "description":      r.description,
        "coaching_cues":    _u(r.coaching_cues),
        "sport_tags":       _u(r.sport_tags),
        "gender_tags":      _u(r.gender_tags),
        "video_url":        r.video_url,
        "notes":            r.notes,
    }

def _fill(row: ExerciseDB, d: ExerciseIn) -> ExerciseDB:
    row.name              = d.name
    row.name_tr           = d.name_tr
    row.movement_pattern  = d.movement_pattern
    row.category          = d.category
    row.equipment         = _j(d.equipment)
    row.difficulty        = d.difficulty
    row.training_quality  = d.training_quality or "Strength"
    row.primary_muscles   = _j(d.primary_muscles)
    row.secondary_muscles = _j(d.secondary_muscles)
    row.description       = d.description
    row.coaching_cues     = _j(d.coaching_cues)
    row.sport_tags        = _j(d.sport_tags)
    row.gender_tags       = _j(d.gender_tags)
    row.video_url         = d.video_url
    row.notes             = d.notes
    return row

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

def _seed_one(db: Session, e: dict):
    row = ExerciseDB(
        name              = e["name"],
        name_tr           = e.get("name_tr"),
        movement_pattern  = e["movement_pattern"],
        category          = e.get("category"),
        equipment         = _j(e.get("equipment", [])),
        difficulty        = e["difficulty"],
        training_quality  = e.get("training_quality", "Strength"),
        primary_muscles   = _j(e.get("primary_muscles", [])),
        secondary_muscles = _j(e.get("secondary_muscles", [])),
        description       = e.get("description"),
        coaching_cues     = _j(e.get("coaching_cues", [])),
        sport_tags        = _j(e.get("sport_tags", [])),
        gender_tags       = _j(e.get("gender_tags", ["all"])),
        video_url         = e.get("video_url"),
        notes             = e.get("notes"),
    )
    db.add(row)

# ── Program helpers ────────────────────────────────────────────────────────────

def _auto_phases(duration_weeks: int, periodization_type: str) -> list[str]:
    phases = []
    for wk in range(1, duration_weeks + 1):
        if periodization_type in ("block", "linear"):
            pct = wk / duration_weeks
            if pct <= 0.50:
                phase = "accumulation"
            elif pct <= 0.875:
                phase = "intensification"
            else:
                phase = "realization"
            if wk % 4 == 0 and wk < duration_weeks:
                phase = "deload"
        else:
            phase = "accumulation"
        phases.append(phase)
    return phases


def _row_program(p: ProgramDB) -> dict:
    return {
        "id":                 p.id,
        "name":               p.name,
        "start_date":         p.start_date,
        "duration_weeks":     p.duration_weeks,
        "days_per_week":      p.days_per_week,
        "goal":               p.goal,
        "periodization_type": p.periodization_type,
        "athlete_level":      p.athlete_level,
        "sport":              p.sport,
        "notes":              p.notes,
        "created_at":         p.created_at,
        "athlete_id":         p.athlete_id,
        "team_id":            p.team_id,
    }


def _row_week(w: ProgramWeekDB) -> dict:
    return {
        "id":          w.id,
        "program_id":  w.program_id,
        "week_number": w.week_number,
        "phase":       w.phase,
        "notes":       w.notes,
    }


def _row_session(s: ProgramSessionDB) -> dict:
    return {
        "id":           s.id,
        "week_id":      s.week_id,
        "day_number":   s.day_number,
        "session_name": s.session_name,
        "sort_order":   s.sort_order,
        "notes":        s.notes,
    }


def _row_one_rm(r: OneRMDB) -> dict:
    return {
        "id":            r.id,
        "exercise_id":   r.exercise_id,
        "exercise_name": r.exercise_name,
        "weight_kg":     r.weight_kg,
        "test_date":     r.test_date,
        "notes":         r.notes,
        "created_at":    r.created_at,
        "athlete_id":    r.athlete_id,
    }


def _row_athlete(a: AthleteDB, db: Session = None) -> dict:
    d = {
        "id":         a.id,
        "user_id":    a.user_id,
        "username":   None,
        "name":       a.name,
        "birth_year": a.birth_year,
        "sport":      a.sport,
        "position":   a.position,
        "gender":     a.gender,
        "level":      a.level,
        "notes":      a.notes,
        "created_at": a.created_at,
    }
    if db:
        d["program_count"] = db.query(ProgramDB).filter(ProgramDB.athlete_id == a.id).count()
        d["one_rm_count"]  = db.query(OneRMDB).filter(OneRMDB.athlete_id == a.id).count()
        if a.user_id:
            u = db.get(UserDB, a.user_id)
            d["username"] = u.username if u else None
    return d


def _row_team(t: TeamDB, db: Session = None) -> dict:
    d = {
        "id":         t.id,
        "name":       t.name,
        "sport":      t.sport,
        "notes":      t.notes,
        "created_at": t.created_at,
    }
    if db:
        d["member_count"]  = db.query(TeamMemberDB).filter(TeamMemberDB.team_id == t.id).count()
        d["program_count"] = db.query(ProgramDB).filter(ProgramDB.team_id == t.id).count()
        members = db.query(TeamMemberDB).filter(TeamMemberDB.team_id == t.id).all()
        athlete_ids = [m.athlete_id for m in members]
        athletes = db.query(AthleteDB).filter(AthleteDB.id.in_(athlete_ids)).order_by(AthleteDB.name).all() if athlete_ids else []
        d["members"] = [{"id": a.id, "name": a.name, "sport": a.sport, "position": a.position, "gender": a.gender, "level": a.level} for a in athletes]
    return d


def _row_guest_athlete(g: GuestAthleteDB) -> dict:
    return {
        "id":         g.id,
        "name":       g.name,
        "country":    g.country,
        "gender":     g.gender,
        "birth_year": g.birth_year,
        "notes":      g.notes,
        "created_at": g.created_at,
    }


def _row_competition_event(ev: CompetitionEventDB) -> dict:
    return {
        "id":                   ev.id,
        "name":                 ev.name,
        "date":                 ev.date,
        "location":             ev.location,
        "level":                ev.level,
        "notes":                ev.notes,
        "created_by_coach_id":  ev.created_by_coach_id,
        "created_at":           ev.created_at,
    }


def _row_competition_result(r: CompetitionResultDB, db: Session) -> dict:
    athlete_name       = None
    guest_athlete_name = None
    event_name         = None
    event_date         = None
    if r.athlete_id:
        a = db.get(AthleteDB, r.athlete_id)
        if a:
            athlete_name = a.name
    if r.guest_athlete_id:
        g = db.get(GuestAthleteDB, r.guest_athlete_id)
        if g:
            guest_athlete_name = g.name
    ev = db.get(CompetitionEventDB, r.competition_event_id)
    if ev:
        event_name = ev.name
        event_date = ev.date
    return {
        "id":                       r.id,
        "competition_event_id":     r.competition_event_id,
        "competition_event_name":   event_name,
        "competition_event_date":   event_date,
        "athlete_id":               r.athlete_id,
        "athlete_name":             athlete_name,
        "guest_athlete_id":         r.guest_athlete_id,
        "guest_athlete_name":       guest_athlete_name,
        "apparatus":                r.apparatus,
        "d_score":                  r.d_score,
        "e_score":                  r.e_score,
        "penalty":                  r.penalty,
        "bonus_point":              r.bonus_point or 0.0,
        "d_score_2":                r.d_score_2,
        "e_score_2":                r.e_score_2,
        "penalty_2":                r.penalty_2,
        "bonus_point_2":            r.bonus_point_2,
        "final_score":              r.final_score,
        "rank":                     r.rank,
        "notes":                    r.notes,
        "created_at":               r.created_at,
    }


def _row_set(s: SessionExerciseSetDB) -> dict:
    return {
        "id":                  s.id,
        "session_exercise_id": s.session_exercise_id,
        "set_number":          s.set_number,
        "reps":                s.reps,
        "intensity_type":      s.intensity_type,
        "intensity_value":     s.intensity_value,
        "notes":               s.notes,
    }


def _row_sex(e: SessionExerciseDB, sets_data: list | None = None) -> dict:
    return {
        "id":              e.id,
        "session_id":      e.session_id,
        "exercise_id":     e.exercise_id,
        "exercise_name":   e.exercise_name,
        "sort_order":      e.sort_order,
        "sets":            e.sets,
        "reps":            e.reps,
        "intensity_type":  e.intensity_type,
        "intensity_value": e.intensity_value,
        "rest_seconds":    e.rest_seconds,
        "tempo":           e.tempo,
        "notes":           e.notes,
        "sets_data":       sets_data if sets_data is not None else [],
        "pairing_group":   e.pairing_group,
        "pairing_slot":    e.pairing_slot,
    }


def _row_sex_full(e: SessionExerciseDB, db: Session) -> dict:
    """Like _row_sex but also fetches per-set data from the DB."""
    sets = (
        db.query(SessionExerciseSetDB)
        .filter(SessionExerciseSetDB.session_exercise_id == e.id)
        .order_by(SessionExerciseSetDB.set_number)
        .all()
    )
    return _row_sex(e, [_row_set(s) for s in sets])

# ── FastAPI app ────────────────────────────────────────────────────────────────

app = FastAPI(title="Exercise Library", version="3.0.0")

# ── ReportLab font kaydı (Türkçe karakter desteği) ───────────────────────────
try:
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    _FONTS_DIR = os.path.join(os.path.dirname(__file__), "fonts")
    pdfmetrics.registerFont(TTFont("DejaVu", os.path.join(_FONTS_DIR, "DejaVuSans.ttf")))
    pdfmetrics.registerFont(TTFont("DejaVu-Bold", os.path.join(_FONTS_DIR, "DejaVuSans-Bold.ttf")))
    _PDF_FONTS_OK = True
except Exception:
    _PDF_FONTS_OK = False


# ── Auth middleware ────────────────────────────────────────────────────────────

@app.middleware("http")
async def auth_middleware(request: Request, call_next):
    path   = request.url.path
    method = request.method

    # Allow non-API routes (static files, root) and the auth endpoints without auth
    if not path.startswith("/api/") or path in (
        "/api/auth/login",
        "/api/auth/signup",
        "/api/auth/google",
        "/api/auth/google-client-id",
    ):
        return await call_next(request)

    # Verify Bearer token
    auth = request.headers.get("authorization", "")
    if not auth.startswith("Bearer "):
        return JSONResponse({"detail": "Not authenticated"}, status_code=401)

    try:
        payload  = _decode_token(auth[7:])
        user_id  = int(payload["sub"])
        role     = payload.get("role", "athlete")
    except Exception:
        return JSONResponse({"detail": "Invalid or expired token"}, status_code=401)

    request.state.user_id   = user_id
    request.state.user_role = role

    # Determine if this path+method requires coach or staff role
    coach_or_staff_only = False
    coach_only_strict    = False  # user management reserved for coach only
    if path.startswith(("/api/exercises", "/api/meta", "/api/seed")):
        coach_or_staff_only = True
    elif path.startswith("/api/programs") and method in ("POST", "PUT", "DELETE"):
        coach_or_staff_only = True
    elif path.startswith(("/api/sessions", "/api/session-exercises")) and method in ("POST", "PUT", "DELETE"):
        coach_or_staff_only = True
    elif path.startswith("/api/one-rm") and method in ("POST", "PUT", "DELETE"):
        coach_or_staff_only = True
    elif path.startswith("/api/users"):
        coach_only_strict = True
    elif path.startswith("/api/athletes") and method in ("POST", "PUT", "DELETE"):
        coach_or_staff_only = True
    elif path.startswith("/api/teams") and method in ("POST", "PUT", "DELETE"):
        coach_or_staff_only = True
    elif path.startswith("/api/guest-athletes") and method in ("POST", "PUT", "DELETE"):
        coach_or_staff_only = True
    elif path.startswith("/api/competition-events") and method in ("POST", "PUT", "DELETE"):
        coach_or_staff_only = True
    elif path.startswith("/api/competition-results") and method in ("POST", "PUT", "DELETE"):
        coach_or_staff_only = True

    if coach_only_strict and role != "coach":
        return JSONResponse({"detail": "Coach access required"}, status_code=403)
    if coach_or_staff_only and role not in ("coach", "staff"):
        return JSONResponse({"detail": "Coach access required"}, status_code=403)

    return await call_next(request)


# ── Auth endpoints ─────────────────────────────────────────────────────────────

@app.post("/api/auth/login")
def login(data: LoginIn, db: Session = Depends(get_db)):
    user = db.query(UserDB).filter(UserDB.username == data.username).first()
    if not user or not user.password_hash or not _verify(data.password, user.password_hash):
        raise HTTPException(401, "Invalid username or password")
    token = _create_token(user.id, user.role)
    athlete = db.query(AthleteDB).filter(AthleteDB.user_id == user.id).first()
    athlete_id = athlete.id if athlete else None
    return {"token": token, "user": {"id": user.id, "username": user.username, "role": user.role, "athlete_id": athlete_id}}


@app.post("/api/auth/signup", status_code=201)
def signup(data: SignupIn, db: Session = Depends(get_db)):
    username = data.username.strip()
    if not username:
        raise HTTPException(400, "Username is required")
    if data.role not in ("athlete", "staff", "coach"):
        raise HTTPException(400, "Role must be 'athlete', 'staff', or 'coach'")
    if len(data.password) < 6:
        raise HTTPException(400, "Password must be at least 6 characters")
    if db.query(UserDB).filter(UserDB.username == username).first():
        raise HTTPException(400, "Username already taken")
    user = UserDB(
        username      = username,
        password_hash = _hash(data.password),
        role          = data.role,
        created_at    = datetime.utcnow().isoformat(),
    )
    db.add(user)
    db.flush()
    if data.role == "athlete":
        db.add(AthleteDB(
            user_id    = user.id,
            name       = username,
            created_at = datetime.utcnow().isoformat(),
        ))
    db.commit()
    db.refresh(user)
    token      = _create_token(user.id, user.role)
    athlete    = db.query(AthleteDB).filter(AthleteDB.user_id == user.id).first()
    athlete_id = athlete.id if athlete else None
    return {"token": token, "user": {"id": user.id, "username": user.username, "role": user.role, "athlete_id": athlete_id}}


@app.get("/api/auth/google-client-id")
def google_client_id():
    return {"client_id": GOOGLE_CLIENT_ID or None}


@app.post("/api/auth/google")
def google_login(data: GoogleLoginIn, db: Session = Depends(get_db)):
    if not GOOGLE_CLIENT_ID:
        raise HTTPException(501, "Google sign-in is not configured on this server")
    try:
        from google.oauth2 import id_token as google_id_token
        from google.auth.transport import requests as google_requests
        idinfo = google_id_token.verify_oauth2_token(
            data.credential,
            google_requests.Request(),
            GOOGLE_CLIENT_ID,
        )
    except Exception:
        raise HTTPException(401, "Invalid Google token")

    gid   = idinfo["sub"]
    email = idinfo.get("email", "")
    name  = idinfo.get("name", "") or email.split("@")[0]

    user = db.query(UserDB).filter(UserDB.google_id == gid).first()
    if not user:
        # Fall back to matching by email for users who registered manually
        if email:
            user = db.query(UserDB).filter(UserDB.email == email).first()
        if user:
            user.google_id = gid
            db.commit()
        else:
            # Create a new account
            base = re.sub(r"[^a-z0-9_]", "", email.split("@")[0].lower()) or "user"
            username = base
            counter  = 1
            while db.query(UserDB).filter(UserDB.username == username).first():
                username = f"{base}{counter}"
                counter += 1
            user = UserDB(
                username      = username,
                password_hash = None,
                role          = "athlete",
                google_id     = gid,
                email         = email,
                created_at    = datetime.utcnow().isoformat(),
            )
            db.add(user)
            db.flush()
            db.add(AthleteDB(
                user_id    = user.id,
                name       = name,
                created_at = datetime.utcnow().isoformat(),
            ))
            db.commit()
            db.refresh(user)

    token = _create_token(user.id, user.role)
    athlete = db.query(AthleteDB).filter(AthleteDB.user_id == user.id).first()
    athlete_id = athlete.id if athlete else None
    return {"token": token, "user": {"id": user.id, "username": user.username, "role": user.role, "athlete_id": athlete_id}}


@app.get("/api/auth/me")
def me(request: Request, db: Session = Depends(get_db)):
    user = db.get(UserDB, request.state.user_id)
    if not user:
        raise HTTPException(404, "User not found")
    athlete = db.query(AthleteDB).filter(AthleteDB.user_id == user.id).first()
    athlete_id = athlete.id if athlete else None
    return {"id": user.id, "username": user.username, "role": user.role, "athlete_id": athlete_id}


# ── User management (coach only) ───────────────────────────────────────────────

@app.get("/api/users")
def list_users(db: Session = Depends(get_db)):
    users = db.query(UserDB).order_by(UserDB.username).all()
    return [{"id": u.id, "username": u.username, "role": u.role, "created_at": u.created_at} for u in users]


@app.post("/api/users", status_code=201)
def create_user(data: UserCreateIn, db: Session = Depends(get_db)):
    if data.role not in ("coach", "staff", "athlete"):
        raise HTTPException(400, "Role must be 'coach', 'staff', or 'athlete'")
    if db.query(UserDB).filter(UserDB.username == data.username).first():
        raise HTTPException(400, "Username already exists")
    user = UserDB(
        username      = data.username,
        password_hash = _hash(data.password),
        role          = data.role,
        created_at    = datetime.utcnow().isoformat(),
    )
    db.add(user)
    db.flush()
    if data.role == "athlete":
        db.add(AthleteDB(
            user_id    = user.id,
            name       = user.username,
            created_at = datetime.utcnow().isoformat(),
        ))
    db.commit()
    db.refresh(user)
    return {"id": user.id, "username": user.username, "role": user.role, "created_at": user.created_at}


@app.delete("/api/users/{uid}", status_code=204)
def delete_user(uid: int, request: Request, db: Session = Depends(get_db)):
    if uid == request.state.user_id:
        raise HTTPException(400, "Cannot delete your own account")
    user = db.get(UserDB, uid)
    if not user:
        raise HTTPException(404, "User not found")
    db.delete(user)
    db.commit()


@app.put("/api/users/{uid}/password", status_code=204)
def change_password(uid: int, data: dict, request: Request, db: Session = Depends(get_db)):
    # Coach can change any user's password; user can change their own
    if uid != request.state.user_id and request.state.user_role != "coach":
        raise HTTPException(403, "Not allowed")
    user = db.get(UserDB, uid)
    if not user:
        raise HTTPException(404, "User not found")
    new_pw = data.get("password", "")
    if not new_pw or len(new_pw) < 4:
        raise HTTPException(400, "Password must be at least 4 characters")
    user.password_hash = _hash(new_pw)
    db.commit()


# ── Exercises API ──────────────────────────────────────────────────────────────

@app.get("/api/exercises")
def list_exercises(
    search:           Optional[str] = Query(None),
    movement_pattern: Optional[str] = Query(None),
    difficulty:       Optional[str] = Query(None),
    training_quality: Optional[str] = Query(None),
    sport_tag:        Optional[str] = Query(None),
    gender_tag:       Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    q = db.query(ExerciseDB)
    if search:
        t = f"%{search}%"
        q = q.filter(or_(
            ExerciseDB.name.ilike(t),
            ExerciseDB.name_tr.ilike(t),
            ExerciseDB.description.ilike(t),
            ExerciseDB.primary_muscles.ilike(t),
            ExerciseDB.secondary_muscles.ilike(t),
        ))
    if movement_pattern:
        q = q.filter(ExerciseDB.movement_pattern == movement_pattern)
    if difficulty:
        q = q.filter(ExerciseDB.difficulty == difficulty)
    if training_quality:
        q = q.filter(ExerciseDB.training_quality == training_quality)
    if sport_tag:
        q = q.filter(ExerciseDB.sport_tags.contains(sport_tag))
    if gender_tag and gender_tag != "all":
        q = q.filter(or_(
            ExerciseDB.gender_tags.contains(gender_tag),
            ExerciseDB.gender_tags.contains("all"),
        ))
    return [_row(r) for r in q.order_by(ExerciseDB.movement_pattern, ExerciseDB.name).all()]


@app.get("/api/exercises/{eid}")
def get_exercise(eid: int, db: Session = Depends(get_db)):
    r = db.get(ExerciseDB, eid)
    if not r:
        raise HTTPException(404, "Exercise not found")
    return _row(r)


@app.post("/api/exercises", status_code=201)
def create_exercise(data: ExerciseIn, db: Session = Depends(get_db)):
    r = _fill(ExerciseDB(), data)
    db.add(r)
    db.commit()
    db.refresh(r)
    return _row(r)


@app.put("/api/exercises/{eid}")
def update_exercise(eid: int, data: ExerciseIn, db: Session = Depends(get_db)):
    r = db.get(ExerciseDB, eid)
    if not r:
        raise HTTPException(404, "Exercise not found")
    _fill(r, data)
    db.commit()
    db.refresh(r)
    return _row(r)


@app.delete("/api/exercises/{eid}", status_code=204)
def delete_exercise(eid: int, db: Session = Depends(get_db)):
    r = db.get(ExerciseDB, eid)
    if not r:
        raise HTTPException(404, "Exercise not found")
    db.delete(r)
    db.commit()


@app.get("/api/meta")
def meta(db: Session = Depends(get_db)):
    rows      = db.query(ExerciseDB).all()
    qualities = sorted({r.training_quality for r in rows if r.training_quality})
    custom    = db.query(MovementPatternDB).order_by(MovementPatternDB.name).all()
    custom_list = [{"name": c.name, "color": c.color} for c in custom]
    all_patterns = sorted(
        {r.movement_pattern for r in rows} | {c["name"] for c in custom_list}
    )
    return {
        "movement_patterns": all_patterns,
        "training_qualities": qualities,
        "custom_patterns": custom_list,
    }


@app.post("/api/meta/patterns", status_code=201)
def add_pattern(data: dict, request: Request, db: Session = Depends(get_db)):
    name = (data.get("name") or "").strip()
    if not name:
        raise HTTPException(400, "Name is required")
    if db.query(MovementPatternDB).filter(MovementPatternDB.name == name).first():
        raise HTTPException(409, "Pattern already exists")
    p = MovementPatternDB(name=name, color=data.get("color") or None)
    db.add(p)
    db.commit()
    return {"name": p.name, "color": p.color}


@app.delete("/api/meta/patterns/{name}", status_code=204)
def delete_pattern(name: str, db: Session = Depends(get_db)):
    from urllib.parse import unquote
    name = unquote(name)
    if db.query(ExerciseDB).filter(ExerciseDB.movement_pattern == name).first():
        raise HTTPException(400, "Cannot delete: exercises are using this pattern")
    p = db.query(MovementPatternDB).filter(MovementPatternDB.name == name).first()
    if not p:
        raise HTTPException(404, "Pattern not found")
    db.delete(p)
    db.commit()


@app.post("/api/seed", status_code=201)
def seed(db: Session = Depends(get_db)):
    if db.query(ExerciseDB).count() > 0:
        return {"status": "already seeded"}
    from seed_data import EXERCISES
    for e in EXERCISES:
        _seed_one(db, e)
    db.commit()
    return {"status": "seeded", "count": len(EXERCISES)}


# ── Programs API ───────────────────────────────────────────────────────────────

@app.get("/api/programs")
def list_programs(request: Request, db: Session = Depends(get_db)):
    if request.state.user_role == "athlete":
        athlete = db.query(AthleteDB).filter(AthleteDB.user_id == request.state.user_id).first()
        if not athlete:
            return []
        team_ids = [m.team_id for m in db.query(TeamMemberDB).filter(TeamMemberDB.athlete_id == athlete.id).all()]
        filters = [ProgramDB.athlete_id == athlete.id]
        if team_ids:
            filters.append(ProgramDB.team_id.in_(team_ids))
        rows = db.query(ProgramDB).filter(or_(*filters)).order_by(ProgramDB.created_at.desc()).all()
    else:
        rows = db.query(ProgramDB).order_by(ProgramDB.created_at.desc()).all()
    result = []
    for p in rows:
        d = _row_program(p)
        weeks = db.query(ProgramWeekDB).filter(ProgramWeekDB.program_id == p.id).order_by(ProgramWeekDB.week_number).all()
        d["phases"] = [w.phase for w in weeks]
        result.append(d)
    return result


@app.post("/api/programs", status_code=201)
def create_program(data: ProgramIn, db: Session = Depends(get_db)):
    p = ProgramDB(
        name               = data.name,
        start_date         = data.start_date,
        duration_weeks     = data.duration_weeks,
        days_per_week      = data.days_per_week,
        goal               = data.goal,
        periodization_type = data.periodization_type,
        athlete_level      = data.athlete_level,
        sport              = data.sport,
        athlete_id         = data.athlete_id,
        team_id            = data.team_id,
        notes              = data.notes,
        created_at         = datetime.utcnow().isoformat(),
    )
    db.add(p)
    db.flush()
    phases = _auto_phases(data.duration_weeks, data.periodization_type)
    for i, phase in enumerate(phases, start=1):
        db.add(ProgramWeekDB(program_id=p.id, week_number=i, phase=phase))
    db.commit()
    db.refresh(p)
    return _row_program(p)


@app.get("/api/programs/{pid}")
def get_program(pid: int, db: Session = Depends(get_db)):
    p = db.get(ProgramDB, pid)
    if not p:
        raise HTTPException(404, "Program not found")
    return _row_program(p)


@app.put("/api/programs/{pid}")
def update_program(pid: int, data: ProgramIn, db: Session = Depends(get_db)):
    p = db.get(ProgramDB, pid)
    if not p:
        raise HTTPException(404, "Program not found")
    p.name               = data.name
    p.start_date         = data.start_date
    p.duration_weeks     = data.duration_weeks
    p.days_per_week      = data.days_per_week
    p.goal               = data.goal
    p.periodization_type = data.periodization_type
    p.athlete_level      = data.athlete_level
    p.sport              = data.sport
    p.notes              = data.notes
    p.team_id            = data.team_id
    db.commit()
    db.refresh(p)
    return _row_program(p)


@app.delete("/api/programs/{pid}", status_code=204)
def delete_program(pid: int, db: Session = Depends(get_db)):
    p = db.get(ProgramDB, pid)
    if not p:
        raise HTTPException(404, "Program not found")
    weeks = db.query(ProgramWeekDB).filter(ProgramWeekDB.program_id == pid).all()
    for w in weeks:
        sessions = db.query(ProgramSessionDB).filter(ProgramSessionDB.week_id == w.id).all()
        for s in sessions:
            for e in db.query(SessionExerciseDB).filter(SessionExerciseDB.session_id == s.id).all():
                db.query(SessionExerciseSetDB).filter(
                    SessionExerciseSetDB.session_exercise_id == e.id
                ).delete()
            db.query(SessionExerciseDB).filter(SessionExerciseDB.session_id == s.id).delete()
            db.delete(s)
        db.delete(w)
    db.delete(p)
    db.commit()


@app.get("/api/programs/{pid}/full")
def get_program_full(pid: int, db: Session = Depends(get_db)):
    p = db.get(ProgramDB, pid)
    if not p:
        raise HTTPException(404, "Program not found")
    result = _row_program(p)
    weeks = db.query(ProgramWeekDB).filter(ProgramWeekDB.program_id == pid).order_by(ProgramWeekDB.week_number).all()
    result["weeks"] = []
    for w in weeks:
        wd = _row_week(w)
        sessions = db.query(ProgramSessionDB).filter(ProgramSessionDB.week_id == w.id).order_by(ProgramSessionDB.day_number, ProgramSessionDB.sort_order).all()
        wd["sessions"] = []
        for s in sessions:
            sd = _row_session(s)
            exercises = db.query(SessionExerciseDB).filter(SessionExerciseDB.session_id == s.id).order_by(SessionExerciseDB.sort_order).all()
            ex_ids = [e.exercise_id for e in exercises if e.exercise_id]
            video_map = {}
            if ex_ids:
                for row in db.query(ExerciseDB.id, ExerciseDB.video_url).filter(ExerciseDB.id.in_(ex_ids)).all():
                    video_map[row.id] = row.video_url
            # Batch-fetch per-set data for all exercises in this session
            ex_ids_in_session = [e.id for e in exercises]
            sets_map: dict[int, list] = {eid: [] for eid in ex_ids_in_session}
            if ex_ids_in_session:
                for s in (
                    db.query(SessionExerciseSetDB)
                    .filter(SessionExerciseSetDB.session_exercise_id.in_(ex_ids_in_session))
                    .order_by(SessionExerciseSetDB.set_number)
                    .all()
                ):
                    sets_map[s.session_exercise_id].append(_row_set(s))

            rows = []
            for e in exercises:
                d = _row_sex(e, sets_map.get(e.id, []))
                d["video_url"] = video_map.get(e.exercise_id)
                rows.append(d)
            sd["exercises"] = rows
            wd["sessions"].append(sd)
        result["weeks"].append(wd)
    return result


@app.get("/api/programs/{pid}/export/excel")
def export_program_excel(pid: int, db: Session = Depends(get_db)):
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    p = db.get(ProgramDB, pid)
    if not p:
        raise HTTPException(404, "Program not found")

    full = get_program_full(pid, db)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Program"

    COLOR_TITLE   = "1E3A5F"
    COLOR_WEEK    = "2D5986"
    COLOR_SESSION = "4A7CF6"
    COLOR_HEADER  = "DBEAFE"

    COLS = ["#", "Egzersiz", "Set", "Tekrar", "Yoğunluk Tipi", "Yoğunluk", "Dinlenme (sn)", "Tempo", "Notlar"]
    COL_WIDTHS = [4, 32, 6, 8, 16, 12, 14, 10, 28]
    N_COLS = len(COLS)

    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    def merge_write(row, text, fill_hex, font_color="FFFFFF", bold=True, font_size=11):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=N_COLS)
        cell = ws.cell(row=row, column=1, value=text)
        cell.fill = PatternFill("solid", fgColor=fill_hex)
        cell.font = Font(bold=bold, color=font_color, size=font_size)
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    thin = Side(border_style="thin", color="D1D5DB")
    row_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    cur = 1

    merge_write(cur, full["name"], COLOR_TITLE, font_size=13)
    ws.row_dimensions[cur].height = 22
    cur += 1

    meta_parts = [
        f"Hedef: {full.get('goal', '')}",
        f"Hafta: {full.get('duration_weeks', '')}",
        f"Gün/Hafta: {full.get('days_per_week', '')}",
        f"Seviye: {full.get('athlete_level', '')}",
        f"Tip: {full.get('periodization_type', '')}",
    ]
    if full.get("start_date"):
        meta_parts.append(f"Başlangıç: {full['start_date']}")
    merge_write(cur, "  ·  ".join(meta_parts), "EFF6FF", font_color="374151", bold=False, font_size=9)
    cur += 2

    for week in full.get("weeks", []):
        phase_label = f" — {week['phase']}" if week.get("phase") else ""
        merge_write(cur, f"  HAFTA {week['week_number']}{phase_label}", COLOR_WEEK, font_size=10)
        ws.row_dimensions[cur].height = 18
        cur += 1

        for session in week.get("sessions", []):
            sess_label = f"  Gün {session['day_number']}: {session.get('session_name') or ''}"
            merge_write(cur, sess_label, COLOR_SESSION, font_size=9)
            ws.row_dimensions[cur].height = 16
            cur += 1

            for c_idx, col_name in enumerate(COLS, 1):
                cell = ws.cell(row=cur, column=c_idx, value=col_name)
                cell.fill = PatternFill("solid", fgColor=COLOR_HEADER)
                cell.font = Font(bold=True, color="1E3A5F", size=8)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = row_border
            ws.row_dimensions[cur].height = 15
            cur += 1

            for ex_idx, ex in enumerate(session.get("exercises", [])):
                # Eşleştirme etiketi: A1, B2 vb. yoksa sıra numarası
                pg = ex.get("pairing_group") or ""
                ps = ex.get("pairing_slot") or ""
                ex_label = f"{pg}{ps}" if pg else str(ex_idx + 1)

                sets_data = ex.get("sets_data") or []
                if sets_data:
                    for s in sets_data:
                        row_vals = [
                            f"{ex_label}-{s.get('set_number', '')}",
                            ex.get("exercise_name", ""),
                            "",
                            s.get("reps", ex.get("reps", "")),
                            s.get("intensity_type", ""),
                            s.get("intensity_value", ""),
                            ex.get("rest_seconds", "") if s.get("set_number") == 1 else "",
                            ex.get("tempo", "") if s.get("set_number") == 1 else "",
                            s.get("notes", "") or (ex.get("notes", "") if s.get("set_number") == 1 else ""),
                        ]
                        for c_idx, val in enumerate(row_vals, 1):
                            cell = ws.cell(row=cur, column=c_idx, value=val if val is not None else "")
                            cell.border = row_border
                            cell.font = Font(size=8)
                            if c_idx == N_COLS:
                                cell.alignment = Alignment(wrap_text=True)
                        cur += 1
                else:
                    row_vals = [
                        ex_label,
                        ex.get("exercise_name", ""),
                        ex.get("sets", ""),
                        ex.get("reps", ""),
                        ex.get("intensity_type", ""),
                        ex.get("intensity_value", ""),
                        ex.get("rest_seconds", ""),
                        ex.get("tempo", ""),
                        ex.get("notes", ""),
                    ]
                    for c_idx, val in enumerate(row_vals, 1):
                        cell = ws.cell(row=cur, column=c_idx, value=val if val is not None else "")
                        cell.border = row_border
                        cell.font = Font(size=8)
                        if c_idx == N_COLS:
                            cell.alignment = Alignment(wrap_text=True)
                    cur += 1

            cur += 1
        cur += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_name = unicodedata.normalize("NFKD", full["name"]).encode("ascii", "ignore").decode("ascii")
    safe_name = re.sub(r"[^\w\s-]", "", safe_name).strip().replace(" ", "_")[:40] or "program"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="program_{pid}_{safe_name}.xlsx"'},
    )


@app.get("/api/programs/{pid}/export/pdf")
def export_program_pdf(pid: int, db: Session = Depends(get_db)):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    p = db.get(ProgramDB, pid)
    if not p:
        raise HTTPException(404, "Program not found")

    full = get_program_full(pid, db)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=1.5 * cm, rightMargin=1.5 * cm,
        topMargin=1.5 * cm, bottomMargin=1.5 * cm,
    )

    base_font = "DejaVu" if _PDF_FONTS_OK else "Helvetica"
    bold_font = "DejaVu-Bold" if _PDF_FONTS_OK else "Helvetica-Bold"

    style_title   = ParagraphStyle("PTitle",   fontName=bold_font, fontSize=16, textColor=colors.HexColor("#1E3A5F"), spaceAfter=4)
    style_meta    = ParagraphStyle("PMeta",    fontName=base_font, fontSize=8,  textColor=colors.HexColor("#6B7280"), spaceAfter=8)
    style_week    = ParagraphStyle("PWeek",    fontName=bold_font, fontSize=11, textColor=colors.white, backColor=colors.HexColor("#2D5986"), spaceBefore=8, spaceAfter=2, leftIndent=4)
    style_session = ParagraphStyle("PSession", fontName=bold_font, fontSize=9,  textColor=colors.white, backColor=colors.HexColor("#4A7CF6"), spaceBefore=4, spaceAfter=2, leftIndent=8)
    style_cell    = ParagraphStyle("PCell",    fontName=base_font, fontSize=7)

    TABLE_COLS = ["#", "Egzersiz", "Set", "Tekrar", "Yoğunluk\nTipi", "Yoğunluk", "Dinl.(sn)", "Notlar"]
    COL_WIDTHS = [0.8 * cm, 5.5 * cm, 1 * cm, 1.5 * cm, 2.5 * cm, 2 * cm, 2 * cm, 3.5 * cm]

    HDR_STYLE = [
        ("BACKGROUND",    (0, 0), (-1, 0), colors.HexColor("#DBEAFE")),
        ("TEXTCOLOR",     (0, 0), (-1, 0), colors.HexColor("#1E3A5F")),
        ("FONTNAME",      (0, 0), (-1, 0), bold_font),
        ("FONTSIZE",      (0, 0), (-1, -1), 7),
        ("FONTNAME",      (0, 1), (-1, -1), base_font),
        ("GRID",          (0, 0), (-1, -1), 0.4, colors.HexColor("#D1D5DB")),
        ("ROWBACKGROUNDS",(0, 1), (-1, -1), [colors.white, colors.HexColor("#F9FAFB")]),
        ("VALIGN",        (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING",   (0, 0), (-1, -1), 3),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 3),
        ("TOPPADDING",    (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]

    flowables = []
    flowables.append(Paragraph(full["name"], style_title))

    meta_parts = [
        f"Hedef: {full.get('goal', '')}",
        f"Hafta: {full.get('duration_weeks', '')}",
        f"Gün/Hafta: {full.get('days_per_week', '')}",
        f"Seviye: {full.get('athlete_level', '')}",
        f"Tip: {full.get('periodization_type', '')}",
    ]
    if full.get("start_date"):
        meta_parts.append(f"Başlangıç: {full['start_date']}")
    flowables.append(Paragraph("  ·  ".join(meta_parts), style_meta))

    for week in full.get("weeks", []):
        phase_label = f" — {week['phase']}" if week.get("phase") else ""
        flowables.append(Paragraph(f"Hafta {week['week_number']}{phase_label}", style_week))

        for session in week.get("sessions", []):
            sess_label = f"Gün {session['day_number']}: {session.get('session_name') or ''}"
            flowables.append(Paragraph(sess_label, style_session))

            data = [TABLE_COLS]
            for ex_idx, ex in enumerate(session.get("exercises", [])):
                # Eşleştirme etiketi: A1, B2 vb. yoksa sıra numarası
                pg = ex.get("pairing_group") or ""
                ps = ex.get("pairing_slot") or ""
                ex_label = f"{pg}{ps}" if pg else str(ex_idx + 1)

                sets_data = ex.get("sets_data") or []
                if sets_data:
                    for s in sets_data:
                        data.append([
                            f"{ex_label}-{s.get('set_number', '')}",
                            Paragraph(ex.get("exercise_name", ""), style_cell),
                            "",
                            str(s.get("reps", ex.get("reps", "")) or ""),
                            str(s.get("intensity_type", "") or ""),
                            str(s.get("intensity_value", "") or ""),
                            str(ex.get("rest_seconds", "") or "") if s.get("set_number") == 1 else "",
                            Paragraph(str(s.get("notes") or (ex.get("notes") if s.get("set_number") == 1 else "") or ""), style_cell),
                        ])
                else:
                    data.append([
                        ex_label,
                        Paragraph(ex.get("exercise_name", ""), style_cell),
                        str(ex.get("sets", "") or ""),
                        str(ex.get("reps", "") or ""),
                        str(ex.get("intensity_type", "") or ""),
                        str(ex.get("intensity_value", "") or ""),
                        str(ex.get("rest_seconds", "") or ""),
                        Paragraph(str(ex.get("notes", "") or ""), style_cell),
                    ])

            if len(data) > 1:
                t = Table(data, colWidths=COL_WIDTHS, repeatRows=1)
                t.setStyle(TableStyle(HDR_STYLE))
                flowables.append(t)
            flowables.append(Spacer(1, 0.2 * cm))

    doc.build(flowables)
    buf.seek(0)

    safe_name = unicodedata.normalize("NFKD", full["name"]).encode("ascii", "ignore").decode("ascii")
    safe_name = re.sub(r"[^\w\s-]", "", safe_name).strip().replace(" ", "_")[:40] or "program"

    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="program_{pid}_{safe_name}.pdf"'},
    )


@app.put("/api/programs/{pid}/weeks/{wnum}/phase")
def update_week_phase(pid: int, wnum: int, data: WeekPhaseIn, db: Session = Depends(get_db)):
    w = db.query(ProgramWeekDB).filter(ProgramWeekDB.program_id == pid, ProgramWeekDB.week_number == wnum).first()
    if not w:
        raise HTTPException(404, "Week not found")
    w.phase = data.phase
    db.commit()
    return _row_week(w)


@app.post("/api/programs/{pid}/sessions", status_code=201)
def add_session(pid: int, data: SessionIn, db: Session = Depends(get_db)):
    p = db.get(ProgramDB, pid)
    if not p:
        raise HTTPException(404, "Program not found")
    w = db.query(ProgramWeekDB).filter(ProgramWeekDB.program_id == pid, ProgramWeekDB.week_number == data.week_number).first()
    if not w:
        raise HTTPException(404, "Week not found")
    existing = db.query(ProgramSessionDB).filter(ProgramSessionDB.week_id == w.id).count()
    s = ProgramSessionDB(
        week_id      = w.id,
        day_number   = data.day_number,
        session_name = data.session_name,
        sort_order   = existing,
        notes        = data.notes,
    )
    db.add(s)
    db.commit()
    db.refresh(s)
    return _row_session(s)


@app.post("/api/programs/{pid}/sync-week1", status_code=200)
def sync_week1_to_all(pid: int, day_number: int = Query(...), db: Session = Depends(get_db)):
    """Copy sessions and exercises from Week 1 for the given day to every other week."""
    week1 = db.query(ProgramWeekDB).filter(
        ProgramWeekDB.program_id == pid, ProgramWeekDB.week_number == 1
    ).first()
    if not week1:
        raise HTTPException(404, "Week 1 not found")

    w1_sessions = db.query(ProgramSessionDB).filter(
        ProgramSessionDB.week_id == week1.id,
        ProgramSessionDB.day_number == day_number,
    ).order_by(ProgramSessionDB.sort_order).all()

    other_weeks = db.query(ProgramWeekDB).filter(
        ProgramWeekDB.program_id == pid, ProgramWeekDB.week_number != 1
    ).all()

    for week in other_weeks:
        # Remove only sessions on this day in the other week
        for old_s in db.query(ProgramSessionDB).filter(
            ProgramSessionDB.week_id == week.id,
            ProgramSessionDB.day_number == day_number,
        ).all():
            for old_e in db.query(SessionExerciseDB).filter(
                SessionExerciseDB.session_id == old_s.id
            ).all():
                db.query(SessionExerciseSetDB).filter(
                    SessionExerciseSetDB.session_exercise_id == old_e.id
                ).delete()
            db.query(SessionExerciseDB).filter(SessionExerciseDB.session_id == old_s.id).delete()
            db.delete(old_s)

        # Copy each Week 1 session and its exercises (including per-set data)
        for s in w1_sessions:
            new_s = ProgramSessionDB(
                week_id      = week.id,
                day_number   = s.day_number,
                session_name = s.session_name,
                sort_order   = s.sort_order,
                notes        = s.notes,
            )
            db.add(new_s)
            db.flush()
            for e in db.query(SessionExerciseDB).filter(
                SessionExerciseDB.session_id == s.id
            ).order_by(SessionExerciseDB.sort_order).all():
                new_e = SessionExerciseDB(
                    session_id      = new_s.id,
                    exercise_id     = e.exercise_id,
                    exercise_name   = e.exercise_name,
                    sort_order      = e.sort_order,
                    sets            = e.sets,
                    reps            = e.reps,
                    intensity_type  = e.intensity_type,
                    intensity_value = e.intensity_value,
                    rest_seconds    = e.rest_seconds,
                    tempo           = e.tempo,
                    notes           = e.notes,
                )
                db.add(new_e)
                db.flush()
                for set_row in db.query(SessionExerciseSetDB).filter(
                    SessionExerciseSetDB.session_exercise_id == e.id
                ).order_by(SessionExerciseSetDB.set_number).all():
                    db.add(SessionExerciseSetDB(
                        session_exercise_id = new_e.id,
                        set_number          = set_row.set_number,
                        reps                = set_row.reps,
                        intensity_type      = set_row.intensity_type,
                        intensity_value     = set_row.intensity_value,
                        notes               = set_row.notes,
                    ))

    db.commit()
    return {"ok": True}


@app.put("/api/sessions/{sid}")
def update_session(sid: int, data: SessionUpdateIn, db: Session = Depends(get_db)):
    s = db.get(ProgramSessionDB, sid)
    if not s:
        raise HTTPException(404, "Session not found")
    if data.session_name is not None:
        s.session_name = data.session_name
    if data.notes is not None:
        s.notes = data.notes
    db.commit()
    db.refresh(s)
    return _row_session(s)


@app.delete("/api/sessions/{sid}", status_code=204)
def delete_session(sid: int, db: Session = Depends(get_db)):
    s = db.get(ProgramSessionDB, sid)
    if not s:
        raise HTTPException(404, "Session not found")
    for e in db.query(SessionExerciseDB).filter(SessionExerciseDB.session_id == sid).all():
        db.query(SessionExerciseSetDB).filter(
            SessionExerciseSetDB.session_exercise_id == e.id
        ).delete()
    db.query(SessionExerciseDB).filter(SessionExerciseDB.session_id == sid).delete()
    db.delete(s)
    db.commit()


def _upsert_sets(exercise_id: int, sets_data: list, db: Session) -> None:
    """Replace all per-set rows for an exercise with the provided sets_data list."""
    db.query(SessionExerciseSetDB).filter(
        SessionExerciseSetDB.session_exercise_id == exercise_id
    ).delete()
    for s in sets_data:
        db.add(SessionExerciseSetDB(
            session_exercise_id = exercise_id,
            set_number          = s.set_number,
            reps                = s.reps,
            intensity_type      = s.intensity_type,
            intensity_value     = s.intensity_value,
            notes               = s.notes,
        ))


@app.post("/api/sessions/{sid}/exercises", status_code=201)
def add_session_exercise(sid: int, data: SessionExerciseIn, db: Session = Depends(get_db)):
    s = db.get(ProgramSessionDB, sid)
    if not s:
        raise HTTPException(404, "Session not found")
    existing = db.query(SessionExerciseDB).filter(SessionExerciseDB.session_id == sid).count()
    ex_name = data.exercise_name
    if data.exercise_id and not ex_name:
        ex = db.get(ExerciseDB, data.exercise_id)
        if ex:
            ex_name = ex.name
    e = SessionExerciseDB(
        session_id      = sid,
        exercise_id     = data.exercise_id,
        exercise_name   = ex_name,
        sort_order      = existing,
        sets            = data.sets,
        reps            = data.reps,
        intensity_type  = data.intensity_type,
        intensity_value = data.intensity_value,
        rest_seconds    = data.rest_seconds,
        tempo           = data.tempo,
        notes           = data.notes,
        pairing_group   = data.pairing_group or None,
        pairing_slot    = data.pairing_slot,
    )
    db.add(e)
    db.flush()   # get e.id before commit
    if data.sets_data:
        _upsert_sets(e.id, data.sets_data, db)
    db.commit()
    db.refresh(e)
    return _row_sex_full(e, db)


@app.put("/api/session-exercises/{seid}")
def update_session_exercise(seid: int, data: SessionExerciseIn, db: Session = Depends(get_db)):
    e = db.get(SessionExerciseDB, seid)
    if not e:
        raise HTTPException(404, "Session exercise not found")
    if data.exercise_id is not None:
        e.exercise_id = data.exercise_id
    if data.exercise_name is not None:
        e.exercise_name = data.exercise_name
    if data.sets is not None:
        e.sets = data.sets
    if data.reps is not None:
        e.reps = data.reps
    if data.intensity_type is not None:
        e.intensity_type = data.intensity_type
    if data.intensity_value is not None:
        e.intensity_value = data.intensity_value
    if data.rest_seconds is not None:
        e.rest_seconds = data.rest_seconds
    if data.tempo is not None:
        e.tempo = data.tempo
    if data.notes is not None:
        e.notes = data.notes
    # pairing_group: boş string gönderilince None'a çevir (grubu silmek için)
    if "pairing_group" in data.model_fields_set:
        e.pairing_group = data.pairing_group or None
    if "pairing_slot" in data.model_fields_set:
        e.pairing_slot = data.pairing_slot
    if data.sets_data is not None:
        _upsert_sets(e.id, data.sets_data, db)
    db.commit()
    db.refresh(e)
    return _row_sex_full(e, db)


@app.delete("/api/session-exercises/{seid}", status_code=204)
def delete_session_exercise(seid: int, db: Session = Depends(get_db)):
    e = db.get(SessionExerciseDB, seid)
    if not e:
        raise HTTPException(404, "Session exercise not found")
    db.query(SessionExerciseSetDB).filter(
        SessionExerciseSetDB.session_exercise_id == seid
    ).delete()
    db.delete(e)
    db.commit()


# ── Per-set CRUD ────────────────────────────────────────────────────────────────

@app.post("/api/session-exercises/{seid}/sets", status_code=201)
def add_exercise_set(seid: int, data: ExerciseSetIn, db: Session = Depends(get_db)):
    e = db.get(SessionExerciseDB, seid)
    if not e:
        raise HTTPException(404, "Session exercise not found")
    s = SessionExerciseSetDB(
        session_exercise_id = seid,
        set_number          = data.set_number,
        reps                = data.reps,
        intensity_type      = data.intensity_type,
        intensity_value     = data.intensity_value,
        notes               = data.notes,
    )
    db.add(s)
    db.commit()
    db.refresh(s)
    return _row_set(s)


@app.put("/api/exercise-sets/{setid}")
def update_exercise_set(setid: int, data: ExerciseSetUpdateIn, db: Session = Depends(get_db)):
    s = db.get(SessionExerciseSetDB, setid)
    if not s:
        raise HTTPException(404, "Set not found")
    if data.reps is not None:
        s.reps = data.reps
    if data.intensity_type is not None:
        s.intensity_type = data.intensity_type
    if data.intensity_value is not None:
        s.intensity_value = data.intensity_value
    if data.notes is not None:
        s.notes = data.notes
    db.commit()
    db.refresh(s)
    return _row_set(s)


@app.delete("/api/exercise-sets/{setid}", status_code=204)
def delete_exercise_set(setid: int, db: Session = Depends(get_db)):
    s = db.get(SessionExerciseSetDB, setid)
    if not s:
        raise HTTPException(404, "Set not found")
    db.delete(s)
    db.commit()

# ── 1RM Records API ────────────────────────────────────────────────────────────

@app.get("/api/one-rm/derived")
def list_derived_one_rm(request: Request, db: Session = Depends(get_db)):
    """Return estimated 1RMs for exercises that have a defined relationship to a
    base exercise with a known 1RM record.  Exercises that already have their own
    direct 1RM record are excluded from the result."""
    # Build a map: UPPER(exercise_name) → latest weight as float
    if request.state.user_role == "athlete":
        athlete = db.query(AthleteDB).filter(AthleteDB.user_id == request.state.user_id).first()
        if not athlete:
            return []
        orm_query = db.query(OneRMDB).filter(OneRMDB.athlete_id == athlete.id)
    else:
        orm_query = db.query(OneRMDB)
    all_records = orm_query.order_by(
        OneRMDB.exercise_name, OneRMDB.test_date.desc(), OneRMDB.created_at.desc()
    ).all()
    orm_map: dict[str, float] = {}
    for rec in all_records:
        key = rec.exercise_name.strip().upper()
        if key not in orm_map:           # keep only the most recent per exercise
            try:
                orm_map[key] = float(rec.weight_kg)
            except ValueError:
                pass

    relationships = db.query(ExerciseRMRelationshipDB).all()

    derived = []
    for rel in relationships:
        ex_key     = rel.exercise_name.strip().upper()
        src_key    = rel.source_name.strip().upper()

        # Skip if the exercise already has its own direct 1RM record
        if ex_key in orm_map:
            continue

        # Skip if we don't know the source's 1RM
        if src_key not in orm_map:
            continue

        source_weight = orm_map[src_key]
        derived_weight = source_weight * rel.factor
        if rel.per_hand:
            derived_weight = derived_weight / 2   # per-dumbbell weight

        pct_display = f"{rel.factor * 100:.3g}%"
        formula = (
            f"{pct_display} of {rel.source_name}"
            + (" / 2 (each hand)" if rel.per_hand else "")
        )

        derived.append({
            "exercise_name":  rel.exercise_name,
            "weight_kg":      round(derived_weight, 2),
            "source_name":    rel.source_name,
            "source_weight":  source_weight,
            "factor":         rel.factor,
            "per_hand":       bool(rel.per_hand),
            "formula":        formula,
        })

    derived.sort(key=lambda x: x["exercise_name"])
    return derived


@app.get("/api/one-rm")
def list_one_rm(
    request: Request,
    search: Optional[str] = Query(None),
    athlete_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
):
    q = db.query(OneRMDB)
    if request.state.user_role == "athlete":
        athlete = db.query(AthleteDB).filter(AthleteDB.user_id == request.state.user_id).first()
        if not athlete:
            return []
        q = q.filter(OneRMDB.athlete_id == athlete.id)
    elif athlete_id:
        q = q.filter(OneRMDB.athlete_id == athlete_id)
    if search:
        q = q.filter(OneRMDB.exercise_name.ilike(f"%{search}%"))
    rows = q.order_by(OneRMDB.exercise_name, OneRMDB.test_date.desc()).all()
    return [_row_one_rm(r) for r in rows]


@app.get("/api/one-rm/exercise/{exercise_id}")
def get_one_rm_for_exercise(exercise_id: int, athlete_id: Optional[int] = Query(None), db: Session = Depends(get_db)):
    q = db.query(OneRMDB).filter(OneRMDB.exercise_id == exercise_id)
    if athlete_id:
        q = q.filter(OneRMDB.athlete_id == athlete_id)
    r = q.order_by(OneRMDB.test_date.desc(), OneRMDB.created_at.desc()).first()
    if not r:
        raise HTTPException(404, "No 1RM record found for this exercise")
    return _row_one_rm(r)


@app.post("/api/one-rm", status_code=201)
def create_one_rm(data: OneRMIn, db: Session = Depends(get_db)):
    name = data.exercise_name
    if not name and data.exercise_id:
        ex = db.get(ExerciseDB, data.exercise_id)
        if ex:
            name = ex.name
    r = OneRMDB(
        exercise_id   = data.exercise_id,
        exercise_name = name or "Unknown",
        weight_kg     = data.weight_kg,
        test_date     = data.test_date,
        notes         = data.notes,
        created_at    = datetime.utcnow().isoformat(),
        athlete_id    = data.athlete_id,
    )
    db.add(r)
    db.commit()
    db.refresh(r)
    return _row_one_rm(r)


@app.put("/api/one-rm/{rid}")
def update_one_rm(rid: int, data: OneRMIn, db: Session = Depends(get_db)):
    r = db.get(OneRMDB, rid)
    if not r:
        raise HTTPException(404, "Record not found")
    r.exercise_id   = data.exercise_id
    r.exercise_name = data.exercise_name
    r.weight_kg     = data.weight_kg
    r.test_date     = data.test_date
    r.notes         = data.notes
    db.commit()
    db.refresh(r)
    return _row_one_rm(r)


@app.delete("/api/one-rm/{rid}", status_code=204)
def delete_one_rm(rid: int, db: Session = Depends(get_db)):
    r = db.get(OneRMDB, rid)
    if not r:
        raise HTTPException(404, "Record not found")
    db.delete(r)
    db.commit()

# ── Athletes API ───────────────────────────────────────────────────────────────

@app.get("/api/athletes")
def list_athletes(db: Session = Depends(get_db)):
    athletes = db.query(AthleteDB).order_by(AthleteDB.name).all()
    return [_row_athlete(a, db) for a in athletes]


@app.post("/api/athletes", status_code=201)
def create_athlete(data: AthleteIn, db: Session = Depends(get_db)):
    # Auto-create a login account from the athlete's name + supplied password
    user_id = None
    if data.password:
        if len(data.password) < 4:
            raise HTTPException(400, "Password must be at least 4 characters")
        username = _athlete_username(data.name, db)
        user = UserDB(
            username      = username,
            password_hash = _hash(data.password),
            role          = "athlete",
            created_at    = datetime.utcnow().isoformat(),
        )
        db.add(user)
        db.flush()          # obtain user.id before committing
        user_id = user.id

    a = AthleteDB(
        user_id    = user_id,
        name       = data.name,
        birth_year = data.birth_year,
        sport      = data.sport,
        position   = data.position,
        gender     = data.gender,
        level      = data.level,
        notes      = data.notes,
        created_at = datetime.utcnow().isoformat(),
    )
    db.add(a)
    db.commit()
    db.refresh(a)
    return _row_athlete(a, db)


@app.get("/api/athletes/{aid}")
def get_athlete(aid: int, db: Session = Depends(get_db)):
    a = db.get(AthleteDB, aid)
    if not a:
        raise HTTPException(404, "Athlete not found")
    return _row_athlete(a, db)


@app.put("/api/athletes/{aid}")
def update_athlete(aid: int, data: AthleteIn, db: Session = Depends(get_db)):
    a = db.get(AthleteDB, aid)
    if not a:
        raise HTTPException(404, "Athlete not found")

    # If a new password is provided, update or create the linked account
    if data.password:
        if len(data.password) < 4:
            raise HTTPException(400, "Password must be at least 4 characters")
        if a.user_id:
            user = db.get(UserDB, a.user_id)
            if user:
                user.password_hash = _hash(data.password)
        else:
            # Create a brand-new account for this athlete
            username = _athlete_username(data.name, db)
            user = UserDB(
                username      = username,
                password_hash = _hash(data.password),
                role          = "athlete",
                created_at    = datetime.utcnow().isoformat(),
            )
            db.add(user)
            db.flush()
            a.user_id = user.id

    a.name       = data.name
    a.birth_year = data.birth_year
    a.sport      = data.sport
    a.position   = data.position
    a.gender     = data.gender
    a.level      = data.level
    a.notes      = data.notes
    db.commit()
    db.refresh(a)
    return _row_athlete(a, db)


@app.delete("/api/athletes/{aid}", status_code=204)
def delete_athlete(aid: int, db: Session = Depends(get_db)):
    a = db.get(AthleteDB, aid)
    if not a:
        raise HTTPException(404, "Athlete not found")
    db.delete(a)
    db.commit()


@app.get("/api/athletes/{aid}/programs")
def get_athlete_programs(aid: int, db: Session = Depends(get_db)):
    a = db.get(AthleteDB, aid)
    if not a:
        raise HTTPException(404, "Athlete not found")
    rows = db.query(ProgramDB).filter(ProgramDB.athlete_id == aid).order_by(ProgramDB.created_at.desc()).all()
    result = []
    for p in rows:
        d = _row_program(p)
        weeks = db.query(ProgramWeekDB).filter(ProgramWeekDB.program_id == p.id).order_by(ProgramWeekDB.week_number).all()
        d["phases"] = [w.phase for w in weeks]
        result.append(d)
    return result


@app.get("/api/athletes/{aid}/one-rm")
def get_athlete_one_rm(aid: int, request: Request, search: Optional[str] = Query(None), db: Session = Depends(get_db)):
    if request.state.user_role == "athlete":
        own = db.query(AthleteDB).filter(AthleteDB.user_id == request.state.user_id).first()
        if not own or own.id != aid:
            raise HTTPException(403, "Access denied")
    a = db.get(AthleteDB, aid)
    if not a:
        raise HTTPException(404, "Athlete not found")
    q = db.query(OneRMDB).filter(OneRMDB.athlete_id == aid)
    if search:
        q = q.filter(OneRMDB.exercise_name.ilike(f"%{search}%"))
    rows = q.order_by(OneRMDB.exercise_name, OneRMDB.test_date.desc()).all()
    return [_row_one_rm(r) for r in rows]


@app.get("/api/athletes/{aid}/one-rm/derived")
def get_athlete_derived_one_rm(aid: int, db: Session = Depends(get_db)):
    """Return estimated 1RMs for this athlete using exercise relationships."""
    all_records = db.query(OneRMDB).filter(OneRMDB.athlete_id == aid).order_by(
        OneRMDB.exercise_name, OneRMDB.test_date.desc(), OneRMDB.created_at.desc()
    ).all()
    orm_map: dict[str, float] = {}
    for rec in all_records:
        key = rec.exercise_name.strip().upper()
        if key not in orm_map:
            try:
                orm_map[key] = float(rec.weight_kg)
            except ValueError:
                pass

    relationships = db.query(ExerciseRMRelationshipDB).all()
    derived = []
    for rel in relationships:
        ex_key  = rel.exercise_name.strip().upper()
        src_key = rel.source_name.strip().upper()
        if ex_key in orm_map or src_key not in orm_map:
            continue
        source_weight  = orm_map[src_key]
        derived_weight = source_weight * rel.factor
        if rel.per_hand:
            derived_weight /= 2
        pct_display = f"{rel.factor * 100:.3g}%"
        formula = f"{pct_display} of {rel.source_name}" + (" / 2 (each hand)" if rel.per_hand else "")
        derived.append({
            "exercise_name": rel.exercise_name,
            "weight_kg":     round(derived_weight, 2),
            "source_name":   rel.source_name,
            "source_weight": source_weight,
            "factor":        rel.factor,
            "per_hand":      bool(rel.per_hand),
            "formula":       formula,
        })
    derived.sort(key=lambda x: x["exercise_name"])
    return derived


# ── Teams ─────────────────────────────────────────────────────────────────────

@app.get("/api/teams")
def list_teams(db: Session = Depends(get_db)):
    teams = db.query(TeamDB).order_by(TeamDB.name).all()
    return [_row_team(t, db) for t in teams]


@app.post("/api/teams", status_code=201)
def create_team(data: TeamIn, db: Session = Depends(get_db)):
    t = TeamDB(
        name       = data.name,
        sport      = data.sport,
        notes      = data.notes,
        created_at = datetime.utcnow().isoformat(),
    )
    db.add(t)
    db.commit()
    db.refresh(t)
    return _row_team(t, db)


@app.get("/api/teams/{tid}")
def get_team(tid: int, db: Session = Depends(get_db)):
    t = db.get(TeamDB, tid)
    if not t:
        raise HTTPException(404, "Team not found")
    return _row_team(t, db)


@app.put("/api/teams/{tid}")
def update_team(tid: int, data: TeamIn, db: Session = Depends(get_db)):
    t = db.get(TeamDB, tid)
    if not t:
        raise HTTPException(404, "Team not found")
    t.name  = data.name
    t.sport = data.sport
    t.notes = data.notes
    db.commit()
    db.refresh(t)
    return _row_team(t, db)


@app.delete("/api/teams/{tid}", status_code=204)
def delete_team(tid: int, db: Session = Depends(get_db)):
    t = db.get(TeamDB, tid)
    if not t:
        raise HTTPException(404, "Team not found")
    db.query(TeamMemberDB).filter(TeamMemberDB.team_id == tid).delete()
    db.delete(t)
    db.commit()


@app.post("/api/teams/{tid}/members", status_code=201)
def add_team_member(tid: int, data: TeamMemberIn, db: Session = Depends(get_db)):
    t = db.get(TeamDB, tid)
    if not t:
        raise HTTPException(404, "Team not found")
    a = db.get(AthleteDB, data.athlete_id)
    if not a:
        raise HTTPException(404, "Athlete not found")
    existing = db.query(TeamMemberDB).filter(
        TeamMemberDB.team_id == tid, TeamMemberDB.athlete_id == data.athlete_id
    ).first()
    if existing:
        raise HTTPException(400, "Athlete is already in this team")
    m = TeamMemberDB(team_id=tid, athlete_id=data.athlete_id)
    db.add(m)
    db.commit()
    return _row_team(t, db)


@app.delete("/api/teams/{tid}/members/{aid}", status_code=204)
def remove_team_member(tid: int, aid: int, db: Session = Depends(get_db)):
    m = db.query(TeamMemberDB).filter(
        TeamMemberDB.team_id == tid, TeamMemberDB.athlete_id == aid
    ).first()
    if not m:
        raise HTTPException(404, "Member not found")
    db.delete(m)
    db.commit()


@app.get("/api/teams/{tid}/programs")
def get_team_programs(tid: int, db: Session = Depends(get_db)):
    t = db.get(TeamDB, tid)
    if not t:
        raise HTTPException(404, "Team not found")
    rows = db.query(ProgramDB).filter(ProgramDB.team_id == tid).order_by(ProgramDB.created_at.desc()).all()
    result = []
    for p in rows:
        d = _row_program(p)
        weeks = db.query(ProgramWeekDB).filter(ProgramWeekDB.program_id == p.id).order_by(ProgramWeekDB.week_number).all()
        d["phases"] = [w.phase for w in weeks]
        result.append(d)
    return result


# ── Guest Athletes ─────────────────────────────────────────────────────────────

@app.get("/api/guest-athletes")
def list_guest_athletes(db: Session = Depends(get_db)):
    rows = db.query(GuestAthleteDB).order_by(GuestAthleteDB.name).all()
    return [_row_guest_athlete(g) for g in rows]


@app.post("/api/guest-athletes", status_code=201)
def create_guest_athlete(data: GuestAthleteIn, db: Session = Depends(get_db)):
    if data.gender not in ("male", "female"):
        raise HTTPException(400, "gender must be 'male' or 'female'")
    g = GuestAthleteDB(
        name       = data.name,
        country    = data.country,
        gender     = data.gender,
        birth_year = data.birth_year,
        notes      = data.notes,
        created_at = datetime.utcnow().isoformat(),
    )
    db.add(g); db.commit(); db.refresh(g)
    return _row_guest_athlete(g)


@app.put("/api/guest-athletes/{gid}")
def update_guest_athlete(gid: int, data: GuestAthleteIn, db: Session = Depends(get_db)):
    g = db.get(GuestAthleteDB, gid)
    if not g:
        raise HTTPException(404, "Guest athlete not found")
    if data.gender not in ("male", "female"):
        raise HTTPException(400, "gender must be 'male' or 'female'")
    g.name       = data.name
    g.country    = data.country
    g.gender     = data.gender
    g.birth_year = data.birth_year
    g.notes      = data.notes
    db.commit(); db.refresh(g)
    return _row_guest_athlete(g)


@app.delete("/api/guest-athletes/{gid}", status_code=204)
def delete_guest_athlete(gid: int, db: Session = Depends(get_db)):
    g = db.get(GuestAthleteDB, gid)
    if not g:
        raise HTTPException(404, "Guest athlete not found")
    db.delete(g); db.commit()


# ── Competition Events ─────────────────────────────────────────────────────────

@app.get("/api/competition-events")
def list_competition_events(db: Session = Depends(get_db)):
    rows = db.query(CompetitionEventDB).order_by(CompetitionEventDB.date.desc()).all()
    return [_row_competition_event(ev) for ev in rows]


@app.post("/api/competition-events", status_code=201)
def create_competition_event(data: CompetitionEventIn, request: Request, db: Session = Depends(get_db)):
    ev = CompetitionEventDB(
        name                = data.name,
        date                = data.date,
        location            = data.location,
        level               = data.level,
        notes               = data.notes,
        created_by_coach_id = request.state.user_id,
        created_at          = datetime.utcnow().isoformat(),
    )
    db.add(ev); db.commit(); db.refresh(ev)
    return _row_competition_event(ev)


@app.put("/api/competition-events/{ceid}")
def update_competition_event(ceid: int, data: CompetitionEventIn, db: Session = Depends(get_db)):
    ev = db.get(CompetitionEventDB, ceid)
    if not ev:
        raise HTTPException(404, "Competition event not found")
    ev.name     = data.name
    ev.date     = data.date
    ev.location = data.location
    ev.level    = data.level
    ev.notes    = data.notes
    db.commit(); db.refresh(ev)
    return _row_competition_event(ev)


@app.delete("/api/competition-events/{ceid}", status_code=204)
def delete_competition_event(ceid: int, db: Session = Depends(get_db)):
    ev = db.get(CompetitionEventDB, ceid)
    if not ev:
        raise HTTPException(404, "Competition event not found")
    db.delete(ev); db.commit()


# ── Competition Results ────────────────────────────────────────────────────────

@app.get("/api/competition-results")
def list_competition_results(
    athlete_id:           Optional[int] = Query(None),
    guest_athlete_id:     Optional[int] = Query(None),
    competition_event_id: Optional[int] = Query(None),
    apparatus:            Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    q = db.query(CompetitionResultDB)
    if athlete_id           is not None: q = q.filter(CompetitionResultDB.athlete_id == athlete_id)
    if guest_athlete_id     is not None: q = q.filter(CompetitionResultDB.guest_athlete_id == guest_athlete_id)
    if competition_event_id is not None: q = q.filter(CompetitionResultDB.competition_event_id == competition_event_id)
    if apparatus:                         q = q.filter(CompetitionResultDB.apparatus == apparatus)
    rows = q.order_by(CompetitionResultDB.created_at.desc()).all()
    return [_row_competition_result(r, db) for r in rows]


@app.post("/api/competition-results", status_code=201)
def create_competition_result(data: CompetitionResultIn, db: Session = Depends(get_db)):
    if not data.athlete_id and not data.guest_athlete_id:
        raise HTTPException(400, "Either athlete_id or guest_athlete_id must be provided")
    if data.athlete_id and data.guest_athlete_id:
        raise HTTPException(400, "Only one of athlete_id or guest_athlete_id may be set")
    jump1 = round(data.d_score + data.e_score - data.penalty + data.bonus_point, 3)
    if data.apparatus == "VT" and data.d_score_2 is not None:
        jump2 = round((data.d_score_2 or 0) + (data.e_score_2 or 0) - (data.penalty_2 or 0) + (data.bonus_point_2 or 0), 3)
        final = round((jump1 + jump2) / 2, 3)
    else:
        final = jump1
    r = CompetitionResultDB(
        competition_event_id = data.competition_event_id,
        athlete_id           = data.athlete_id,
        guest_athlete_id     = data.guest_athlete_id,
        apparatus            = data.apparatus,
        d_score              = data.d_score,
        e_score              = data.e_score,
        penalty              = data.penalty,
        bonus_point          = data.bonus_point,
        d_score_2            = data.d_score_2 if data.apparatus == "VT" else None,
        e_score_2            = data.e_score_2 if data.apparatus == "VT" else None,
        penalty_2            = data.penalty_2 if data.apparatus == "VT" else None,
        bonus_point_2        = data.bonus_point_2 if data.apparatus == "VT" else None,
        final_score          = final,
        rank                 = data.rank,
        notes                = data.notes,
        created_at           = datetime.utcnow().isoformat(),
    )
    db.add(r); db.commit(); db.refresh(r)
    return _row_competition_result(r, db)


@app.put("/api/competition-results/{rid}")
def update_competition_result(rid: int, data: CompetitionResultIn, db: Session = Depends(get_db)):
    r = db.get(CompetitionResultDB, rid)
    if not r:
        raise HTTPException(404, "Result not found")
    if not data.athlete_id and not data.guest_athlete_id:
        raise HTTPException(400, "Either athlete_id or guest_athlete_id must be provided")
    if data.athlete_id and data.guest_athlete_id:
        raise HTTPException(400, "Only one of athlete_id or guest_athlete_id may be set")
    r.competition_event_id = data.competition_event_id
    r.athlete_id           = data.athlete_id
    r.guest_athlete_id     = data.guest_athlete_id
    r.apparatus            = data.apparatus
    r.d_score              = data.d_score
    r.e_score              = data.e_score
    r.penalty              = data.penalty
    r.bonus_point          = data.bonus_point
    r.d_score_2            = data.d_score_2 if data.apparatus == "VT" else None
    r.e_score_2            = data.e_score_2 if data.apparatus == "VT" else None
    r.penalty_2            = data.penalty_2 if data.apparatus == "VT" else None
    r.bonus_point_2        = data.bonus_point_2 if data.apparatus == "VT" else None
    jump1 = round(data.d_score + data.e_score - data.penalty + data.bonus_point, 3)
    if data.apparatus == "VT" and data.d_score_2 is not None:
        jump2 = round((data.d_score_2 or 0) + (data.e_score_2 or 0) - (data.penalty_2 or 0) + (data.bonus_point_2 or 0), 3)
        r.final_score = round((jump1 + jump2) / 2, 3)
    else:
        r.final_score = jump1
    r.rank                 = data.rank
    r.notes                = data.notes
    db.commit(); db.refresh(r)
    return _row_competition_result(r, db)


@app.delete("/api/competition-results/{rid}", status_code=204)
def delete_competition_result(rid: int, db: Session = Depends(get_db)):
    r = db.get(CompetitionResultDB, rid)
    if not r:
        raise HTTPException(404, "Result not found")
    db.delete(r); db.commit()


@app.get("/api/athletes/{aid}/competition-results")
def get_athlete_competition_results(aid: int, db: Session = Depends(get_db)):
    rows = db.query(CompetitionResultDB).filter(
        CompetitionResultDB.athlete_id == aid
    ).order_by(CompetitionResultDB.created_at.desc()).all()
    return [_row_competition_result(r, db) for r in rows]


# ── Startup ────────────────────────────────────────────────────────────────────

@app.on_event("startup")
def auto_seed():
    db = SessionLocal()
    try:
        if db.query(ExerciseDB).count() == 0:
            from seed_data import EXERCISES
            for e in EXERCISES:
                _seed_one(db, e)
            db.commit()

        if db.query(ExerciseRMRelationshipDB).count() == 0:
            for ex_name, src_name, factor, per_hand in RM_RELATIONSHIPS_SEED:
                db.add(ExerciseRMRelationshipDB(
                    exercise_name = ex_name,
                    source_name   = src_name,
                    factor        = factor,
                    per_hand      = per_hand,
                ))
            db.commit()

        if db.query(UserDB).count() == 0:
            db.add(UserDB(
                username      = "admin",
                password_hash = _hash("admin1234"),
                role          = "coach",
                created_at    = datetime.utcnow().isoformat(),
            ))
            db.commit()
            print("\n" + "=" * 50)
            print("  Default coach account created:")
            print("    Username : admin")
            print("    Password : admin1234")
            print("  Change this password after first login!")
            print("=" * 50 + "\n")
    finally:
        db.close()


# ── Static files ───────────────────────────────────────────────────────────────

@app.get("/static/{path:path}")
def serve_static(path: str):
    from pathlib import Path
    file_path = Path("static") / path
    if not file_path.exists():
        raise HTTPException(404)
    response = FileResponse(file_path)
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response.headers["Pragma"] = "no-cache"
    return response

@app.get("/")
def root():
    response = FileResponse("static/index.html")
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response.headers["Pragma"] = "no-cache"
    return response
